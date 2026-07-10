# -*- coding: utf-8 -*-
"""Extrai o TEXTO das ArrayFormula (I/AA/Listas) e formulas de ano E/F/G,
alem do cabecalho completo da Base de dados. Saida em .txt UTF-8."""
import os
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

BASE = r"C:\Users\t825026\OneDrive - Santander Office 365\Documentos\DataMaster"
OUT = os.path.join(BASE, "datamaster-app", "_assets_extracted")
TEMPLATE = os.path.join(BASE, "template_plano_de_contas.xlsx")

wb = load_workbook(TEMPLATE, data_only=False)

def get(ws, addr):
    v = ws[addr].value
    if isinstance(v, ArrayFormula):
        return f"ARRAYFORMULA ref={v.ref} text={v.text}"
    return v

lines = []
def w(s=""):
    lines.append(str(s))

shadow = None
listas = None
base = None
for n in wb.sheetnames:
    nl = n.strip().lower()
    if nl.startswith("shadow"):
        shadow = wb[n]
    elif nl == "listas":
        listas = wb[n]
    elif nl.startswith("base"):
        base = wb[n]

w("=== SHADOW year value cols E/F/G (row 5) and H/I ===")
for a in ["B2", "E2", "F2", "G2", "E5", "F5", "G5", "H5", "I5"]:
    w(f"{a}: {get(shadow, a)}")

w("\n=== SHADOW DRE year cols W/X/Y (row 5) and Z/AA/AB ===")
for a in ["T2", "W2", "X2", "Y2", "W5", "X5", "Y5", "Z5", "AA5", "AB5"]:
    w(f"{a}: {get(shadow, a)}")

w("\n=== LISTAS A2/C2/E2 ===")
for a in ["A2", "C2", "E2"]:
    w(f"{a}: {get(listas, a)}")

w("\n=== BASE DE DADOS headers row1 (A..AG) ===")
from openpyxl.utils import get_column_letter
hdrs = []
for ci in range(1, 34):
    col = get_column_letter(ci)
    v = base[f"{col}1"].value
    if v not in (None, ""):
        hdrs.append(f"{col}1={v!r}")
w("\n".join(hdrs))

# extra ranges list for AB inversor
w("\n=== SHADOW AM1/AN1 (inversor list source) ===")
for a in ["AM1", "AN1", "AM2", "AN2"]:
    w(f"{a}: {get(shadow, a)}")

with open(os.path.join(OUT, "formulas.txt"), "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print("wrote formulas.txt", len("\n".join(lines)), "chars")
