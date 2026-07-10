# -*- coding: utf-8 -*-
"""Resumo COMPACTO do template.json: lista de contas (Plano de Contas),
uma formula-exemplo por coluna de formula, cabecalhos e validacoes."""
import json
import os

OUT = os.path.dirname(os.path.abspath(__file__))
tpl = json.load(open(os.path.join(OUT, "template.json"), encoding="utf-8"))

lines = []
def w(s=""):
    lines.append(s)

def cells_of(name):
    for k, sh in tpl["sheets"].items():
        if k.strip().lower() == name:
            return sh.get("full", {}).get("cells", {}), sh
    return {}, {}

def fmt(cells, addr):
    c = cells.get(addr)
    if c is None:
        return None
    return ("[F]" + str(c["v"])) if c.get("t") == "f" else c.get("v")

w("SHEETNAMES: " + repr(tpl["sheetnames"]))
w()

# ---------------- SHADOW ----------------
cells, sh = cells_of("shadow")
w("=" * 70)
w(f"SHADOW state={sh.get('state')} dims={sh.get('dims')}")
w("\n-- HEADERS / MARKERS (rows 1-4, cols A..AB) --")
from openpyxl.utils import get_column_letter, column_index_from_string
for r in range(1, 5):
    parts = []
    for ci in range(1, 37):
        col = get_column_letter(ci)
        v = fmt(cells, f"{col}{r}")
        if v not in (None, ""):
            parts.append(f"{col}{r}={v}")
    if parts:
        w("  " + " | ".join(parts))

w("\n-- CONTAS ATIVO/PASSIVO (col A) + formulas B/H/I (rows 5-80) --")
for r in range(5, 81):
    a = fmt(cells, f"A{r}")
    if a in (None, ""):
        continue
    b = fmt(cells, f"B{r}")
    h = fmt(cells, f"H{r}")
    i = fmt(cells, f"I{r}")
    w(f"  A{r}={a!r}")
    for lbl, val in (("B", b), ("H", h), ("I", i)):
        if val not in (None, ""):
            w(f"      {lbl}{r}={val}")

w("\n-- CONTAS DRE (col S) + formulas T/Z/AA/AB (rows 5-40) --")
for r in range(5, 41):
    s = fmt(cells, f"S{r}")
    if s in (None, ""):
        continue
    t = fmt(cells, f"T{r}")
    z = fmt(cells, f"Z{r}")
    aa = fmt(cells, f"AA{r}")
    ab = fmt(cells, f"AB{r}")
    w(f"  S{r}={s!r}")
    for lbl, val in (("T", t), ("Z", z), ("AA", aa), ("AB", ab)):
        if val not in (None, ""):
            w(f"      {lbl}{r}={val}")

w("\n-- DATA VALIDATIONS (shadow) --")
for dv in sh.get("data_validations", []) or []:
    w(f"  {dv['sqref']} type={dv['type']} allow_blank={dv['allow_blank']} f1={dv['formula1']}")

# ---------------- LISTAS ----------------
cells, sh = cells_of("listas")
w("\n" + "=" * 70)
w(f"LISTAS state={sh.get('state')} dims={sh.get('dims')}")
for r in range(1, 6):
    for col in ("A", "B", "C", "D", "E", "F"):
        v = fmt(cells, f"{col}{r}")
        if v not in (None, ""):
            w(f"  {col}{r}={v}")

# ---------------- RASTREABILIDADE ----------------
cells, sh = cells_of("rastreabilidade")
w("\n" + "=" * 70)
w(f"RASTREABILIDADE state={sh.get('state')} dims={sh.get('dims')}")
hdr = []
for ci in range(1, 20):
    col = get_column_letter(ci)
    v = fmt(cells, f"{col}1")
    if v not in (None, ""):
        hdr.append(f"{col}1={v}")
w("  HEADERS: " + " | ".join(hdr))
# linha 2 (exemplo de formula, se houver)
row2 = []
for ci in range(1, 20):
    col = get_column_letter(ci)
    v = fmt(cells, f"{col}2")
    if v not in (None, ""):
        row2.append(f"{col}2={v}")
if row2:
    w("  ROW2: " + " | ".join(row2))

# ---------------- BASE DE DADOS ----------------
for k, sh in tpl["sheets"].items():
    if k.strip().lower().startswith("base"):
        cells = sh.get("full", {}).get("cells", {})
        w("\n" + "=" * 70)
        w(f"BASE DE DADOS state={sh.get('state')} dims={sh.get('dims')}")
        hdr = []
        for ci in range(1, 40):
            col = get_column_letter(ci)
            v = fmt(cells, f"{col}1")
            if v not in (None, ""):
                hdr.append(f"{col}={v}")
        w("  HEADERS: " + " || ".join(hdr))

txt = "\n".join(lines)
with open(os.path.join(OUT, "template_summary.txt"), "w", encoding="utf-8") as f:
    f.write(txt)
print("wrote template_summary.txt", len(txt), "chars", len(lines), "lines")
