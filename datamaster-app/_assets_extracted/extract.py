# -*- coding: utf-8 -*-
"""Extrai a estrutura dos Excels do CustomGPT para JSON, para servir de
referencia/seed do app web. Acessa por caminho ABSOLUTO (os arquivos do
OneDrive nao aparecem no listdir, mas sao acessiveis pelo path exato)."""
import json
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\t825026\OneDrive - Santander Office 365\Documentos\DataMaster"
OUT = os.path.join(BASE, "datamaster-app", "_assets_extracted")

TEMPLATE = os.path.join(BASE, "template_plano_de_contas.xlsx")
DICIONARIO = os.path.join(BASE, "Dicionário de Contas.xlsx")
SINAL = os.path.join(BASE, "INTERPRETAÇÃO DE SINAL DRE.xlsx")


def cell_repr(cell):
    """Retorna dict com valor e, se for formula, a string da formula."""
    v = cell.value
    if v is None:
        return None
    dt = cell.data_type  # 'f' = formula, 'n' number, 's' string, ...
    return {"v": v if not isinstance(v, (bytes,)) else str(v), "t": dt}


def dump_sheet_full(ws, max_row=None, max_col=None):
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column
    cells = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            rep = cell_repr(cell)
            if rep is not None:
                cells[f"{get_column_letter(c)}{r}"] = rep
    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "dims": ws.dimensions,
        "cells": cells,
    }


def dump_table(ws):
    """Le a sheet como tabela: primeira linha = header, demais = registros."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"title": ws.title, "headers": [], "records": []}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = []
    for r in rows[1:]:
        if all(x is None or str(x).strip() == "" for x in r):
            continue
        rec = {}
        for i, h in enumerate(headers):
            rec[h if h else f"col{i}"] = r[i] if i < len(r) else None
        records.append(rec)
    return {"title": ws.title, "headers": headers, "records": records}


def data_validations(ws):
    out = []
    for dv in getattr(ws.data_validations, "dataValidation", []) or []:
        out.append({
            "sqref": str(getattr(dv, "sqref", "")),
            "type": dv.type,
            "formula1": dv.formula1,
            "allow_blank": dv.allow_blank,
        })
    return out


def main():
    result = {}

    # ---- TEMPLATE ----
    wb = load_workbook(TEMPLATE, data_only=False)
    tpl = {"sheetnames": wb.sheetnames, "sheets": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        nl = name.strip().lower()
        entry = {
            "title": name,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "dims": ws.dimensions,
            "data_validations": data_validations(ws),
        }
        # Shadow e Listas: dump completo (sao a base da logica)
        if nl == "shadow" or nl.startswith("shadow"):
            entry["full"] = dump_sheet_full(ws, max_row=85, max_col=40)
        elif nl == "listas":
            entry["full"] = dump_sheet_full(ws, max_row=10, max_col=8)
        else:
            # demais: so cabecalho (linha 1) + linha 2
            entry["full"] = dump_sheet_full(ws, max_row=3, max_col=min(ws.max_column, 40))
        tpl["sheets"][name] = entry
    result["template"] = tpl
    with open(os.path.join(OUT, "template.json"), "w", encoding="utf-8") as f:
        json.dump(tpl, f, ensure_ascii=False, indent=2, default=str)

    # ---- DICIONARIO ----
    wbd = load_workbook(DICIONARIO, data_only=True)
    dic = {"sheetnames": wbd.sheetnames, "sheets": {}}
    for name in wbd.sheetnames:
        dic["sheets"][name] = dump_table(wbd[name])
    with open(os.path.join(OUT, "dicionario.json"), "w", encoding="utf-8") as f:
        json.dump(dic, f, ensure_ascii=False, indent=2, default=str)

    # ---- SINAL ----
    wbs = load_workbook(SINAL, data_only=True)
    sin = {"sheetnames": wbs.sheetnames, "sheets": {}}
    for name in wbs.sheetnames:
        sin["sheets"][name] = dump_table(wbs[name])
    with open(os.path.join(OUT, "sinal.json"), "w", encoding="utf-8") as f:
        json.dump(sin, f, ensure_ascii=False, indent=2, default=str)

    # ---- Resumo no stdout ----
    print("TEMPLATE sheets:", wb.sheetnames)
    print("DICIONARIO sheets:", wbd.sheetnames,
          "| registros:", {k: len(v["records"]) for k, v in dic["sheets"].items()})
    print("SINAL sheets:", wbs.sheetnames,
          "| registros:", {k: len(v["records"]) for k, v in sin["sheets"].items()})
    print("OK -> JSONs em", OUT)


if __name__ == "__main__":
    main()
