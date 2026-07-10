"""
Pos-processamento idempotente: aplica as 6 validacoes de dados
obrigatorias na aba Shadow do template OCR de BP/DRE.

Uso:
    python aplicar_validacoes.py <caminho_do_xlsx>

Se nenhum caminho for passado, procura o xlsx mais recentemente modificado
no diretorio atual.

Este script e seguro para rodar multiplas vezes (idempotente):
remove qualquer DV pre-existente nas faixas alvo antes de re-aplicar.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


RETIRAR_FORMULA = "listas!$E$2:$E$1048576"
ADICIONAR_FORMULA = "listas!$C$2:$C$1048576"

RETIRAR_RANGES = ["J5:M39", "J45:M79", "AC5:AF39"]
ADICIONAR_RANGES = ["N5:Q39", "N45:Q79", "AG5:AJ39"]


def find_shadow_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == "shadow":
            return wb[name]
    for name in wb.sheetnames:
        ws = wb[name]
        a1 = str(ws["A1"].value or "").strip().lower()
        if a1.startswith("shadow empresa"):
            return ws
    raise RuntimeError(
        "Nao foi possivel localizar a aba Shadow. "
        f"Abas disponiveis: {wb.sheetnames}"
    )


def remove_data_validations_in_ranges(shadow_ws, target_ranges):
    dvs = getattr(shadow_ws.data_validations, "dataValidation", [])
    if not dvs:
        return
    remaining = []
    for dv in list(dvs):
        sqref = str(getattr(dv, "sqref", "") or "")
        should_remove = any(target in sqref for target in target_ranges)
        if not should_remove:
            remaining.append(dv)
    shadow_ws.data_validations.dataValidation = remaining


def apply_shadow_data_validations(shadow_ws) -> None:
    remove_data_validations_in_ranges(
        shadow_ws, RETIRAR_RANGES + ADICIONAR_RANGES
    )
    for cell_range in RETIRAR_RANGES:
        dv = DataValidation(
            type="list",
            formula1=RETIRAR_FORMULA,
            allow_blank=True,
            showDropDown=False,
        )
        shadow_ws.add_data_validation(dv)
        dv.add(cell_range)
    for cell_range in ADICIONAR_RANGES:
        dv = DataValidation(
            type="list",
            formula1=ADICIONAR_FORMULA,
            allow_blank=True,
            showDropDown=False,
        )
        shadow_ws.add_data_validation(dv)
        dv.add(cell_range)


def verify_applied(shadow_ws) -> None:
    expected = {r: RETIRAR_FORMULA for r in RETIRAR_RANGES}
    expected.update({r: ADICIONAR_FORMULA for r in ADICIONAR_RANGES})
    existing = [
        (
            str(getattr(dv, "sqref", "") or ""),
            str(getattr(dv, "formula1", "") or "").lstrip("="),
        )
        for dv in getattr(shadow_ws.data_validations, "dataValidation", [])
    ]
    missing = []
    for target_range, target_formula in expected.items():
        found = any(
            target_range in sqref and formula1 == target_formula
            for sqref, formula1 in existing
        )
        if not found:
            missing.append(target_range)
    if missing:
        raise RuntimeError(
            "Validacoes obrigatorias nao encontradas apos save: "
            + ", ".join(missing)
        )


def resolve_target(args: list[str]) -> Path:
    if args:
        p = Path(args[0])
        if not p.exists():
            raise FileNotFoundError(p)
        return p
    cwd = Path.cwd()
    candidates = sorted(
        cwd.glob("*.xlsx"),
        key=lambda x: x.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            "Nenhum .xlsx encontrado no diretorio atual. "
            "Passe o caminho como argumento."
        )
    return candidates[0]


def main(argv: list[str]) -> int:
    target = resolve_target(argv)
    print(f"Alvo: {target}")

    wb = load_workbook(target)
    shadow_ws = find_shadow_sheet(wb)
    print(f"Shadow: {shadow_ws.title}")

    before = len(list(getattr(shadow_ws.data_validations, "dataValidation", [])))
    apply_shadow_data_validations(shadow_ws)
    after = len(list(getattr(shadow_ws.data_validations, "dataValidation", [])))
    print(f"DVs antes: {before} | depois (em memoria): {after}")

    wb.save(target)

    wb2 = load_workbook(target)
    shadow2 = find_shadow_sheet(wb2)
    verify_applied(shadow2)

    print("OK: 6 validacoes obrigatorias aplicadas e verificadas.")
    for dv in shadow2.data_validations.dataValidation:
        print(f"  {str(dv.sqref)} -> {dv.formula1}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
