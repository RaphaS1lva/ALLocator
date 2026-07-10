from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import time
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from datetime import date, datetime
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Font, PatternFill

REQUIRED_TEMPLATE_SHEETS = {"rastreabilidade", "listas"}

# Coluna "Alocacao da Hierarquia" (Sim/Nao). No template novo ela ocupa a
# posicao D (layout A..N: Origem, Hierarquia, Totalizador=C, Alocacao=D,
# Pagina Ref=E, Ano1..3=F:H, Grupo=I, Sub=J, Destino=K, Tipo=L, Chave=M,
# Chave Destino=N). E AUTORITATIVA: vem do julgamento do modelo (Sim = nivel
# escolhido p/ alocacao; Nao = contexto); ausente => Nao. O Destino e'
# preenchido em TODAS as linhas (sugestao); somente as linhas "Sim" geram
# Chave/Chave Destino (logo so elas contam nos SUMIFS da Shadow).
ALOCACAO_HIER_HEADER = "Alocação da Hierarquia"
ALOCACAO_SIM = "Sim"
ALOCACAO_NAO = "Não"

# Flag de totalizador: origens que aparecem como Hierarquia de alguma abertura
# (pai/totalizador). Agora vai na coluna dedicada "Totalizador" (C, Sim/Nao),
# nao mais como sufixo no nome. O sufixo abaixo e' mantido apenas para limpar,
# por idempotencia, eventuais " - Totalizador" preexistentes no campo Hierarquia.
TOTALIZADOR_SUFFIX = " - Totalizador"

# Guardrail (read-only) de SUB-CONSOLIDACAO. Limiar EDITAVEL: quando um
# totalizador (conta-pai) esta marcado 'Não' mas tem PELO MENOS esta quantidade
# de aberturas atomizadas, TODAS alocadas (Alocação=Sim) para o MESMO destino,
# o parecer sugere PROMOVER o totalizador a 'Sim' e rebaixar as aberturas a
# contexto (Não), reduzindo a fragmentacao da Shadow. Diretriz pratica do
# analista: ate ~4 origens alocadas por hierarquia (Regras §2); a partir de
# 5 ja sinaliza para reavaliar a consolidacao.
TOTALIZADOR_PROMOCAO_MIN_FILHOS = 5

SHADOW_HEADER_MARKERS = {
    "A1": "CNPJ:",
    "A2": "EMPRESA:",
    "A3": "GRUPO:",
    "A4": "AUDITADO:",
    "A5": "CONSOLIDADO:",
}

PL_SPECIFIC_ACCOUNTS = {
    "PARTICIPAÇÕES MINORITÁRIAS",
    "CAPITAL SOCIAL",
    "LUCROS ACUMULADOS",
    "OUTRAS RESERVAS",
}

PL_SPECIFIC_ACCOUNTS_NORM = set()


# =========================================================
# Nome padronizado do arquivo de saida (EDITAVEL pelo time)
# =========================================================
# Ajuste APENAS esta string para mudar a convencao de nome do arquivo gerado.
# Campos disponiveis:
#   {empresa} = nome da empresa (sanitizado: sem acentos/espacos/chars ilegais)
#   {cnpj}    = CNPJ apenas com digitos
#   {data}    = data de hoje no formato ddMMAAAA
#   {ano}     = maior (mais recente) ano reconhecido no documento
# A extensao (.xlsx/.xlsm) e anexada automaticamente conforme o template.
OUTPUT_NAME_PATTERN = "{empresa}_{cnpj}_Output_{data}_ALLOCATOR_{ano}"

# Versao do GPT/knowledge (CustomGPT). Editavel pelo time; pode ser
# sobrescrita via --versao-gpt. Vai para a aba `Base de dados` (col "Versao do GPT").
VERSAO_GPT = "v1.06"

# Formato numerico das colunas de ano (D:F) na Rastreabilidade. Padrao
# contabil: milhar; negativo entre parenteses; zero exibido vazio.
YEAR_VALUE_NUMBER_FORMAT = '#,##0;(#,##0);""'


# =========================================================
# Abas internas adicionadas ao arquivo final (EDITÁVEL)
# =========================================================
# rastreabilidade_inicial = cópia CONGELADA da Rastreabilidade (referência p/
# espelho/diff futuro). "Base de dados" = respostas confirmadas + metadados (aba ja existente no template novo; o script preenche, nao recria).
# Ambas saem OCULTAS e PROTEGIDAS por senha (anti-edição); as DEMAIS abas
# continuam editáveis e visíveis (uso proteção por planilha, SEM trava de
# estrutura do workbook). Lembrete: proteção .xlsx NÃO é criptografia, apenas
# evita edição acidental. Troque a senha via --lock-password ou aqui.
SNAPSHOT_SHEET_NAME = "rastreabilidade_inicial"
BASE_DADOS_SHEET_NAME = "Base de dados"
LOCK_PASSWORD = "ALLocator#2026"
LOCK_SHEET_STATE = "hidden"  # use "veryHidden" para esconder até do menu Reexibir


# =========================================================
# Helpers
# =========================================================
def detect_keep_vba(path: Path) -> bool:
    return path.suffix.lower() == ".xlsm"


def open_workbook(path: Path):
    return load_workbook(path, keep_vba=detect_keep_vba(path), data_only=False)


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def strong_partial_match(a: str, b: str) -> bool:
    if not a or not b:
        return False
    return a == b or a in b or b in a


def tokenize(text: str) -> set[str]:
    return {t for t in normalize_text(text).split() if t}


def token_overlap_score(a: str, b: str) -> float:
    ta = tokenize(a)
    tb = tokenize(b)
    if not ta or not tb:
        return 0.0
    inter = len(ta & tb)
    union = len(ta | tb)
    return inter / union if union else 0.0


def sort_year_key(value: Any):
    text = str(value or "").strip()
    try:
        return (0, int(text))
    except Exception:
        return (1, text)


def try_float(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def coerce_number(value: Any):
    """Converte o valor para int/float quando possivel, para que as colunas de
    ano da Rastreabilidade sejam NUMERICAS (e o SUMIFS da Shadow funcione).
    Aceita numeros e strings tipo '1234', '1.234,56', '1,234.56', '-12',
    '(12)'. Retorna None quando vazio e preserva o texto original se nao for
    numero (nunca quebra o pipeline)."""
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text == "":
        return None
    cleaned = text.replace("R$", "").replace("$", "").replace(" ", "").strip()
    neg = False
    if cleaned.startswith("(") and cleaned.endswith(")"):
        neg = True
        cleaned = cleaned[1:-1]
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    elif "." in cleaned and re.fullmatch(r"\d{1,3}(\.\d{3})+", cleaned):
        # ponto sozinho como separador de milhar PT-BR (ex.: 1.000 -> 1000)
        cleaned = cleaned.replace(".", "")
    try:
        num = float(cleaned)
    except ValueError:
        return value
    if neg:
        num = -num
    return int(num) if num == int(num) else num


def build_header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None:
            headers[str(value).strip()] = col
    return headers


def first_empty_row(ws, start_row: int = 2, key_col: int = 1) -> int:
    row = start_row
    while ws.cell(row=row, column=key_col).value not in (None, ""):
        row += 1
    return row


def make_structural_key(destino: str, grupo: str, sub_categoria: str) -> tuple[str, str, str]:
    return (
        str(destino or "").strip(),
        str(grupo or "").strip(),
        str(sub_categoria or "").strip(),
    )


def clear_column_from_row(ws, col_idx: int, start_row: int = 2) -> None:
    for row in range(start_row, ws.max_row + 1):
        ws.cell(row=row, column=col_idx).value = None


def get_non_empty_values(ws, ranges: list[str]) -> set[str]:
    values = set()
    for cell_range in ranges:
        for row in ws[cell_range]:
            for cell in row:
                value = str(cell.value or "").strip()
                if value:
                    values.add(value)
    return values


def init_normalized_constants() -> None:
    global PL_SPECIFIC_ACCOUNTS_NORM
    if not PL_SPECIFIC_ACCOUNTS_NORM:
        PL_SPECIFIC_ACCOUNTS_NORM = {normalize_text(x) for x in PL_SPECIFIC_ACCOUNTS}


def is_pl_specific_account(value: Any) -> bool:
    init_normalized_constants()
    return normalize_text(value) in PL_SPECIFIC_ACCOUNTS_NORM


def normalize_group(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    norm = normalize_text(text)
    if norm == "ativo":
        return "Ativo"
    if norm == "passivo":
        return "Passivo"
    if norm == "dre":
        return "DRE"
    return text.strip()


def normalize_subcategory(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    norm = normalize_text(text)
    if norm == "circulante":
        return "Circulante"
    if norm == "nao circulante":
        return "Não Circulante"
    if norm == "pl":
        return "PL"
    if norm == "dre":
        return "DRE"
    return text.strip()


def apply_special_classification_rules(row: dict[str, Any]) -> dict[str, Any]:
    item = dict(row)

    origem = str(item.get("origem", "")).strip()
    destino = str(item.get("destino_template", "")).strip()
    grupo = normalize_group(item.get("grupo", ""))
    sub = normalize_subcategory(item.get("sub_categoria", ""))

    if is_pl_specific_account(origem) or is_pl_specific_account(destino):
        item["grupo"] = "Passivo"
        item["sub_categoria"] = "PL"
        return item

    if grupo == "DRE":
        item["grupo"] = "DRE"
        item["sub_categoria"] = "DRE"
        return item

    item["grupo"] = grupo
    item["sub_categoria"] = sub
    return item


def get_sheet_by_name_case_insensitive(wb, expected_name: str):
    expected_norm = str(expected_name).strip().lower()
    for ws in wb.worksheets:
        if ws.title.strip().lower() == expected_norm:
            return ws
    raise ValueError(f"Aba não encontrada: {expected_name}")


# =========================================================
# Shadow selection
# =========================================================
def is_valid_shadow_sheet(ws) -> bool:
    for cell_ref, expected in SHADOW_HEADER_MARKERS.items():
        actual = str(ws[cell_ref].value or "").strip()
        if cell_ref == "A3":
            if normalize_text(actual) != normalize_text(expected):
                return False
        else:
            if actual != expected:
                return False
    return True


def list_shadow_sheets(wb) -> list[dict[str, str]]:
    options: list[dict[str, str]] = []
    for ws in wb.worksheets:
        if is_valid_shadow_sheet(ws):
            company = str(ws["B2"].value or "").strip() or ws.title
            options.append({"sheet_name": ws.title, "company_name": company})
    return options


def resolve_shadow_sheet(wb, shadow_company: str = ""):
    options = list_shadow_sheets(wb)

    if not options:
        raise ValueError(
            "Nenhuma aba válida de Memória Anterior foi encontrada com os marcadores esperados em A1:A5."
        )

    if len(options) == 1:
        return wb[options[0]["sheet_name"]]

    if not str(shadow_company or "").strip():
        formatted = "\n".join(
            f"{idx}. {opt['company_name']} [{opt['sheet_name']}]"
            for idx, opt in enumerate(options, start=1)
        )
        raise ValueError(
            "Foram encontradas múltiplas abas válidas de Memória Anterior. "
            "Informe --shadow-company com o número ou o nome da empresa.\n"
            f"Opções:\n{formatted}"
        )

    choice = str(shadow_company).strip()

    if choice.isdigit():
        idx = int(choice)
        if 1 <= idx <= len(options):
            return wb[options[idx - 1]["sheet_name"]]
        raise ValueError(f"Índice de Shadow inválido: {choice}")

    choice_norm = normalize_text(choice)
    for opt in options:
        if normalize_text(opt["company_name"]) == choice_norm:
            return wb[opt["sheet_name"]]

    raise ValueError(
        f"Não foi possível localizar a Shadow para '{choice}'. "
        "Use o número da opção ou o nome da empresa em B2."
    )


# =========================================================
# Validação da estrutura
# =========================================================
def validate_template(wb, shadow_company: str = ""):
    existing_names = {ws.title.strip().lower() for ws in wb.worksheets}
    missing = sorted(REQUIRED_TEMPLATE_SHEETS - existing_names)
    if missing:
        raise ValueError(f"Abas ausentes no template: {missing}")

    shadow_ws = resolve_shadow_sheet(wb, shadow_company)

    if not is_valid_shadow_sheet(shadow_ws):
        raise ValueError(
            "Aba de Memória Anterior inválida: os marcadores A1:A5 não conferem com o layout esperado."
        )

    return (
        shadow_ws,
        get_sheet_by_name_case_insensitive(wb, "rastreabilidade"),
        get_sheet_by_name_case_insensitive(wb, "listas"),
    )


def validate_rastreabilidade_headers(ws) -> dict[str, int]:
    header_map = build_header_map(ws)
    required = [
        "Origem",
        "Hierarquia",
        "Totalizador",
        ALOCACAO_HIER_HEADER,
        "Página Referência",
        "Ano 1",
        "Ano 2",
        "Ano 3",
        "Grupo",
        "Sub Categoria",
        "Destino no Template",
        "Tipo de Mapeamento",
        "Chave",
        "Chave Destino",
    ]
    missing = [h for h in required if h not in header_map]
    if missing:
        raise ValueError(f"Colunas ausentes na Rastreabilidade: {missing}")
    return header_map


# =========================================================
# OCR JSON
# =========================================================
def load_rows(json_path: Path) -> list[dict[str, Any]]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("O JSON de entrada deve conter uma lista de lançamentos.")
    return data


def normalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        item.setdefault("origem", "")
        item.setdefault("hierarquia", "")
        item.setdefault("pagina_referencia", "")
        item.setdefault("valor", "")
        item.setdefault("grupo", "")
        item.setdefault("sub_categoria", "")
        item.setdefault("destino_template", "")
        item.setdefault("ano", "")
        item.setdefault("tipo_mapeamento", "")
        item.setdefault("alocacao_hierarquia", "")
        item["_origem_norm"] = normalize_text(item.get("origem", ""))

        if not str(item["tipo_mapeamento"]).strip():
            item["tipo_mapeamento"] = "Julgamental"

        item = apply_special_classification_rules(item)
        normalized.append(item)

    return sorted(
        normalized,
        key=lambda r: (sort_year_key(r.get("ano")), str(r.get("origem", "")).lower()),
    )


def normalize_alocacao_value(raw: Any) -> str:
    """Normaliza a 'Alocação da Hierarquia' (Sim/Não).

    AUTORITATIVA: vem do julgamento do modelo (campo 'alocacao_hierarquia').
    'Sim' = nivel escolhido para alocacao (conta na Chave/Chave Destino);
    'Não'/vazio = contexto (Destino no Template e' apenas sugestao, nao conta).
    O Destino e' preenchido em TODAS as linhas, entao NAO se infere mais o
    'Sim' a partir da existencia de destino."""
    if normalize_text(str(raw or "")) in ("sim", "s", "yes", "true", "1"):
        return ALOCACAO_SIM
    return ALOCACAO_NAO


# =========================================================
# Índice estrutural do template
# =========================================================
def infer_subcategoria_from_position(grupo: str, row: int) -> str:
    if grupo == "Ativo":
        return "Circulante" if row <= 22 else "Não Circulante"
    if grupo == "Passivo":
        if row >= 75:
            return "PL"
        return "Circulante" if row <= 60 else "Não Circulante"
    if grupo == "DRE":
        return "DRE"
    return ""


def build_template_account_index(shadow_ws) -> dict[str, list[dict[str, str]]]:
    index: dict[str, list[dict[str, str]]] = defaultdict(list)

    for row in range(5, 40):
        conta = str(shadow_ws[f"A{row}"].value or "").strip()
        if conta:
            index[conta].append(
                {
                    "grupo": "Ativo",
                    "sub_categoria": infer_subcategoria_from_position("Ativo", row),
                    "cell": f"A{row}",
                }
            )

    for row in range(45, 80):
        conta = str(shadow_ws[f"A{row}"].value or "").strip()
        if conta:
            index[conta].append(
                {
                    "grupo": "Passivo",
                    "sub_categoria": infer_subcategoria_from_position("Passivo", row),
                    "cell": f"A{row}",
                }
            )

    for row in range(5, 40):
        conta = str(shadow_ws[f"S{row}"].value or "").strip()
        if conta:
            index[conta].append(
                {
                    "grupo": "DRE",
                    "sub_categoria": "DRE",
                    "cell": f"S{row}",
                }
            )

    return index


def find_duplicate_destination_names(template_index) -> dict[str, list[dict[str, str]]]:
    duplicates = {}
    for destino, items in template_index.items():
        if len(items) > 1:
            duplicates[destino] = items
    return duplicates


def is_destination_compatible(destino, grupo, sub_categoria, template_index) -> bool:
    candidates = template_index.get(str(destino or "").strip(), [])
    if not candidates:
        return False

    grupo = normalize_group(grupo)
    sub_categoria = normalize_subcategory(sub_categoria)

    for item in candidates:
        if item["grupo"] != grupo:
            continue
        if item["sub_categoria"] != sub_categoria:
            continue
        return True
    return False


def enforce_template_classification(rows, template_index):
    cleaned = []
    for idx, row in enumerate(rows, start=1):
        item = apply_special_classification_rules(row)
        destino = str(item.get("destino_template", "")).strip()

        if destino:
            ok = is_destination_compatible(
                destino=destino,
                grupo=item.get("grupo", ""),
                sub_categoria=item.get("sub_categoria", ""),
                template_index=template_index,
            )
            if not ok:
                raise ValueError(
                    f"Alocação estrutural inválida na linha {idx}: "
                    f"origem={item.get('origem')} "
                    f"grupo={item.get('grupo')} "
                    f"sub_categoria={item.get('sub_categoria')} "
                    f"destino={destino}"
                )

        cleaned.append(item)
    return cleaned

# =========================================================
# Ordenacao e merge da Rastreabilidade
# =========================================================
GROUP_SUB_ORDER: dict[tuple[str, str], int] = {
    ("Ativo", "Circulante"): 0,
    ("Ativo", "Não Circulante"): 1,
    ("Passivo", "Circulante"): 2,
    ("Passivo", "Não Circulante"): 3,
    ("Passivo", "PL"): 4,
    ("DRE", "DRE"): 5,
}


def compute_years(rows: list[dict[str, Any]]) -> list[str]:
    """Retorna ate 3 anos distintos ordenados do mais antigo ao mais recente."""
    return sorted(
        {str(r.get("ano", "")).strip() for r in rows if str(r.get("ano", "")).strip()},
        key=sort_year_key,
    )[:3]


def sanitize_filename_part(value: Any) -> str:
    """Normaliza um pedaco de nome de arquivo: remove acentos, caracteres ilegais
    em nome de arquivo (\\ / : * ? \" < > |) e espacos."""
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r'[\\/:*?"<>|]', "", text)
    text = re.sub(r"\s+", "", text)
    return text


def only_digits(value: Any) -> str:
    """Mantem apenas digitos (ex.: CNPJ '12.345.678/0001-90' -> '12345678000190')."""
    return re.sub(r"\D", "", str(value or ""))


def build_output_filename(empresa: Any, cnpj: Any, years, today=None, ext: str = ".xlsx") -> str:
    """
    Constroi o nome padronizado do arquivo de saida a partir de OUTPUT_NAME_PATTERN.
    - empresa: nome da empresa (sera sanitizado)
    - cnpj: CNPJ (sera reduzido a digitos)
    - years: anos reconhecidos; usa o MAIOR (mais recente)
    - today: date opcional (default = hoje), formatado ddMMAAAA
    - ext: extensao do arquivo (.xlsx ou .xlsm)
    Edite OUTPUT_NAME_PATTERN (no topo do modulo) para mudar a convencao.
    """
    d = today or date.today()
    anos = [str(y).strip() for y in (years or []) if str(y).strip()]
    ano = sorted(anos, key=sort_year_key)[-1] if anos else "SemAno"
    nome = OUTPUT_NAME_PATTERN.format(
        empresa=sanitize_filename_part(empresa) or "Empresa",
        cnpj=only_digits(cnpj) or "SemCNPJ",
        data=d.strftime("%d%m%Y"),
        ano=ano,
    )
    if ext and not nome.lower().endswith(ext.lower()):
        nome += ext
    return nome


def build_destination_order(shadow_ws) -> dict[tuple[str, str, str], int]:
    """
    Constroi um dicionario {(grupo, sub_categoria, destino_normalizado): row_idx}
    com a ordem em que cada destino aparece no Plano de Contas do template
    (lido da Shadow). A chave usa o destino normalizado para tolerar diferencas
    de capitalizacao/acentos/whitespace.
    """
    order: dict[tuple[str, str, str], int] = {}

    # Ativo: linhas 5 a 41 na coluna A
    for row in range(5, 42):
        conta = str(shadow_ws[f"A{row}"].value or "").strip()
        if conta:
            sub = infer_subcategoria_from_position("Ativo", row)
            key = ("Ativo", sub, normalize_text(conta))
            order.setdefault(key, row)

    # Passivo + PL: linhas 45 a 84 na coluna A
    for row in range(45, 85):
        conta = str(shadow_ws[f"A{row}"].value or "").strip()
        if conta:
            sub = infer_subcategoria_from_position("Passivo", row)
            key = ("Passivo", sub, normalize_text(conta))
            order.setdefault(key, row)

    # DRE: linhas 5 a 41 na coluna S
    for row in range(5, 42):
        conta = str(shadow_ws[f"S{row}"].value or "").strip()
        if conta:
            key = ("DRE", "DRE", normalize_text(conta))
            order.setdefault(key, row)

    return order


def merge_rows_for_rastreabilidade(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Agrupa rows pela chave estrutural (origem_norm, grupo, sub_categoria,
    destino_template, tipo_mapeamento) e consolida:
      - paginas: concatenacao distinta separada por virgula
      - valores: dict {ano: valor} (ultima ocorrencia ganha em duplicata)

    Para linhas de CONTEXTO (sem destino), a hierarquia (conta-pai) tambem
    entra na chave, para nao colapsar subcontas homonimas de pais distintos
    (ex.: varios "Outros" sob pais diferentes). Linhas ALOCADAS (com destino)
    mantem a chave original.

    Mantem o nome original da origem (primeira ocorrencia).
    """
    merged: dict[tuple, dict[str, Any]] = {}

    for row in rows:
        origem = str(row.get("origem", "")).strip()
        grupo = str(row.get("grupo", "")).strip()
        sub = str(row.get("sub_categoria", "")).strip()
        destino = str(row.get("destino_template", "")).strip()
        tipo = str(row.get("tipo_mapeamento", "")).strip()

        key = (normalize_text(origem), grupo, sub, destino, tipo)
        if not destino:
            # Linhas de contexto (Alocacao=Nao) podem compartilhar nomes
            # genericos sob pais diferentes; incluir a hierarquia evita
            # colapsa-las e perder dado. Linhas com destino nao sao afetadas.
            key = key + (normalize_text(str(row.get("hierarquia", "")).strip()),)

        if key not in merged:
            merged[key] = {
                "origem": origem,
                "hierarquia": str(row.get("hierarquia", "")).strip(),
                "alocacao_hierarquia": str(row.get("alocacao_hierarquia", "")).strip(),
                "_paginas_set": set(),
                "_paginas_list": [],
                "grupo": grupo,
                "sub_categoria": sub,
                "destino_template": destino,
                "tipo_mapeamento": tipo,
                "valores_por_ano": {},
            }

        item = merged[key]

        pag = str(row.get("pagina_referencia", "")).strip()
        if pag and pag not in item["_paginas_set"]:
            item["_paginas_set"].add(pag)
            item["_paginas_list"].append(pag)

        ano = str(row.get("ano", "")).strip()
        if ano:
            item["valores_por_ano"][ano] = row.get("valor")

    result: list[dict[str, Any]] = []
    for item in merged.values():
        item["pagina_referencia"] = ", ".join(item["_paginas_list"])
        item.pop("_paginas_set", None)
        item.pop("_paginas_list", None)
        result.append(item)

    return result


def sort_merged_rows(
    merged: list[dict[str, Any]],
    destination_order: dict[tuple[str, str, str], int],
) -> list[dict[str, Any]]:
    """
    Ordena rows pela ordem do Plano de Contas do template:
      1. Ordem do Grupo/Sub Categoria (Ativo C, Ativo NC, Passivo C, Passivo NC,
         Passivo PL, DRE)
      2. Linhas com destino antes das sem destino (Julgamental sem alocacao)
      3. Posicao do destino na Shadow
      4. Empate: ordem alfabetica de origem
    """
    def key(item: dict[str, Any]):
        grupo = item.get("grupo", "")
        sub = item.get("sub_categoria", "")
        destino = str(item.get("destino_template", "")).strip()

        gs_idx = GROUP_SUB_ORDER.get((grupo, sub), 99)

        if destino:
            d_norm = normalize_text(destino)
            d_idx = destination_order.get((grupo, sub, d_norm), 9999)
            has_destino = 0
        else:
            d_idx = 99999
            has_destino = 1

        return (gs_idx, has_destino, d_idx, str(item.get("origem", "")).lower())

    return sorted(merged, key=key)



# =========================================================
# Dicionário
# =========================================================
def load_dictionary(dictionary_path: Path) -> list[dict[str, str]]:
    wb = open_workbook(dictionary_path)
    # Resolucao de cabecalho tolerante a caixa/acentos: aceita "ORIGEM",
    # "Origem" ou "origem" (normalize_text baixa a caixa e remove acentos).
    required = {"origem", "destino no template", "grupo", "sub categoria"}

    for ws in wb.worksheets:
        raw_headers = build_header_map(ws)
        headers = {normalize_text(name): col for name, col in raw_headers.items()}
        if not required.issubset(headers.keys()):
            continue

        rows: list[dict[str, str]] = []
        for row_idx in range(2, ws.max_row + 1):
            origem = ws.cell(row_idx, headers["origem"]).value
            destino = ws.cell(row_idx, headers["destino no template"]).value
            grupo = ws.cell(row_idx, headers["grupo"]).value
            sub = ws.cell(row_idx, headers["sub categoria"]).value

            if any(v not in (None, "") for v in [origem, destino, grupo, sub]):
                item = {
                    "Origem": str(origem or "").strip(),
                    "Destino no Template": str(destino or "").strip(),
                    "Grupo": normalize_group(grupo),
                    "Sub Categoria": normalize_subcategory(sub),
                    "_origem_norm": normalize_text(origem),
                }
                item = apply_special_classification_rules(
                    {
                        "origem": item["Origem"],
                        "destino_template": item["Destino no Template"],
                        "grupo": item["Grupo"],
                        "sub_categoria": item["Sub Categoria"],
                    }
                )
                rows.append(
                    {
                        "Origem": str(origem or "").strip(),
                        "Destino no Template": str(destino or "").strip(),
                        "Grupo": item["grupo"],
                        "Sub Categoria": item["sub_categoria"],
                        "_origem_norm": normalize_text(origem),
                    }
                )

        if rows:
            return rows

    raise ValueError(
        "Não foi encontrada nenhuma sheet no Dicionário com as colunas: "
        "Origem, Destino no Template, Grupo, Sub Categoria."
    )


def match_dictionary_entry(
    origem: str,
    grupo: str,
    sub_categoria: str,
    dictionary_rows: list[dict[str, str]],
) -> dict[str, str] | None:
    origem_norm = normalize_text(origem)
    grupo = normalize_group(grupo)
    sub_categoria = normalize_subcategory(sub_categoria)

    exact_candidates: list[dict[str, str]] = []
    partial_candidates: list[tuple[float, dict[str, str]]] = []

    for row in dictionary_rows:
        dict_origem_norm = row["_origem_norm"]
        dict_grupo = normalize_group(row["Grupo"])
        dict_sub = normalize_subcategory(row["Sub Categoria"])

        if grupo and dict_grupo and grupo != dict_grupo:
            continue
        if sub_categoria and dict_sub and sub_categoria != dict_sub:
            continue

        if origem_norm == dict_origem_norm:
            exact_candidates.append(row)
            continue

        if strong_partial_match(origem_norm, dict_origem_norm):
            partial_candidates.append((token_overlap_score(origem_norm, dict_origem_norm), row))
            continue

        overlap = token_overlap_score(origem_norm, dict_origem_norm)
        if overlap >= 0.60:
            partial_candidates.append((overlap, row))

    if exact_candidates:
        exact_candidates.sort(
            key=lambda r: (
                0 if normalize_subcategory(r["Sub Categoria"]) == sub_categoria else 1,
                -len(tokenize(r["Origem"])),
                len(r["Origem"]),
            )
        )
        return exact_candidates[0]

    if partial_candidates:
        partial_candidates.sort(
            key=lambda x: (
                -x[0],
                0 if normalize_subcategory(x[1]["Sub Categoria"]) == sub_categoria else 1,
                -len(tokenize(x[1]["Origem"])),
            )
        )
        return partial_candidates[0][1]

    return None


def apply_dictionary_mapping(
    rows: list[dict[str, Any]],
    dictionary_rows: list[dict[str, str]],
) -> list[dict[str, Any]]:
    updated = []

    for item in rows:
        row = apply_special_classification_rules(item)

        if str(row.get("destino_template", "")).strip():
            updated.append(row)
            continue

        match = match_dictionary_entry(
            origem=str(row.get("origem", "")).strip(),
            grupo=str(row.get("grupo", "")).strip(),
            sub_categoria=str(row.get("sub_categoria", "")).strip(),
            dictionary_rows=dictionary_rows,
        )

        if match:
            row["destino_template"] = match["Destino no Template"]
            row["grupo"] = match["Grupo"] or row.get("grupo", "")
            row["sub_categoria"] = match["Sub Categoria"] or row.get("sub_categoria", "")
            row["tipo_mapeamento"] = "Dicionário"
            row = apply_special_classification_rules(row)

        updated.append(row)

    return updated


def annotate_dictionary_source(
    rows: list[dict[str, Any]],
    dictionary_rows: list[dict[str, str]],
) -> tuple[list[dict[str, Any]], int]:
    """Guardrail de rotulagem: quando o modelo ja trouxe a linha com
    destino_template preenchido e tipo_mapeamento ficou 'Julgamental' (ou
    vazio), consulta o dicionario; se a origem casar com o MESMO destino,
    rotula tipo_mapeamento = 'Dicionario'. NUNCA altera o destino, apenas a
    procedencia. Preserva 'Memoria Anterior' e 'Dicionario' ja
    definidos. Retorna (rows, quantidade_rotulada)."""
    relabeled = 0
    for row in rows:
        destino = str(row.get("destino_template", "")).strip()
        if not destino:
            continue
        tipo = str(row.get("tipo_mapeamento", "")).strip()
        if tipo not in ("", "Julgamental"):
            continue
        match = match_dictionary_entry(
            origem=str(row.get("origem", "")).strip(),
            grupo=str(row.get("grupo", "")).strip(),
            sub_categoria=str(row.get("sub_categoria", "")).strip(),
            dictionary_rows=dictionary_rows,
        )
        if match and normalize_text(match["Destino no Template"]) == normalize_text(destino):
            row["tipo_mapeamento"] = "Dicionário"
            relabeled += 1
    return rows, relabeled


# =========================================================
# Shadow / Memória anterior
# =========================================================
def split_memory(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, ArrayFormula):
        return []

    raw = str(value).strip()
    if not raw or raw.startswith("="):
        return []

    upper_raw = raw.upper()
    if "SUMIFS(" in upper_raw or "IFERROR(" in upper_raw or "LET(" in upper_raw:
        return []

    parts = [p.strip() for p in raw.replace(" + ", "+").split("+") if p.strip()]
    cleaned: list[str] = []

    for part in parts:
        item = part.strip().strip("()").strip()
        if item and not item.startswith("="):
            cleaned.append(item)

    return cleaned


def unique_preserve(values: list[str]) -> list[str]:
    seen = set()
    result = []
    for value in values:
        key = normalize_text(value)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value.strip())
    return result


def format_memory(values: list[str]) -> str:
    values = unique_preserve(values)
    return " + ".join(f"({v})" for v in values if v.strip())


def grouped_origins(rows: list[dict[str, Any]]) -> dict[tuple[str, str, str], list[str]]:
    grouped: dict[tuple[str, str, str], list[str]] = defaultdict(list)

    for item in rows:
        destino = str(item.get("destino_template", "")).strip()
        origem = str(item.get("origem", "")).strip()
        grupo = str(item.get("grupo", "")).strip()
        sub = str(item.get("sub_categoria", "")).strip()

        if destino and origem:
            key = make_structural_key(destino, grupo, sub)
            grouped[key].append(f"{origem}|{grupo}|{sub}")

    return grouped


def load_shadow_memory_index(shadow_ws) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []

    for row in range(5, 40):
        destino = str(shadow_ws[f"A{row}"].value or "").strip()
        memoria = split_memory(shadow_ws[f"C{row}"].value)
        grupo = "Ativo"
        sub = infer_subcategoria_from_position("Ativo", row)
        anos = [shadow_ws[f"E{row}"].value, shadow_ws[f"F{row}"].value, shadow_ws[f"G{row}"].value]
        if destino:
            for origem in memoria:
                origem_base = origem.split("|")[0].strip() if "|" in origem else origem
                items.append(
                    {
                        "origem_memoria": origem_base,
                        "_origem_memoria_norm": normalize_text(origem_base),
                        "destino_template": destino,
                        "grupo": grupo,
                        "sub_categoria": sub,
                        "anos_shadow": [str(v).strip() for v in anos if v not in (None, "")],
                    }
                )

    for row in range(45, 80):
        destino = str(shadow_ws[f"A{row}"].value or "").strip()
        memoria = split_memory(shadow_ws[f"C{row}"].value)
        grupo = "Passivo"
        sub = infer_subcategoria_from_position("Passivo", row)
        anos = [shadow_ws[f"E{row}"].value, shadow_ws[f"F{row}"].value, shadow_ws[f"G{row}"].value]
        if destino:
            for origem in memoria:
                origem_base = origem.split("|")[0].strip() if "|" in origem else origem
                items.append(
                    {
                        "origem_memoria": origem_base,
                        "_origem_memoria_norm": normalize_text(origem_base),
                        "destino_template": destino,
                        "grupo": grupo,
                        "sub_categoria": sub,
                        "anos_shadow": [str(v).strip() for v in anos if v not in (None, "")],
                    }
                )

    for row in range(5, 40):
        destino = str(shadow_ws[f"S{row}"].value or "").strip()
        memoria = split_memory(shadow_ws[f"U{row}"].value)
        grupo = "DRE"
        sub = "DRE"
        anos = [shadow_ws[f"W{row}"].value, shadow_ws[f"X{row}"].value, shadow_ws[f"Y{row}"].value]
        if destino:
            for origem in memoria:
                origem_base = origem.split("|")[0].strip() if "|" in origem else origem
                items.append(
                    {
                        "origem_memoria": origem_base,
                        "_origem_memoria_norm": normalize_text(origem_base),
                        "destino_template": destino,
                        "grupo": grupo,
                        "sub_categoria": sub,
                        "anos_shadow": [str(v).strip() for v in anos if v not in (None, "")],
                    }
                )
    return items


def match_shadow_entry(
    origem: str,
    grupo: str,
    sub_categoria: str,
    shadow_items: list[dict[str, str]],
) -> dict[str, str] | None:
    origem_norm = normalize_text(origem)
    grupo = normalize_group(grupo)
    sub_categoria = normalize_subcategory(sub_categoria)

    exact: list[dict[str, str]] = []
    partial: list[tuple[float, dict[str, str]]] = []

    for item in shadow_items:
        if grupo and item["grupo"] and grupo != item["grupo"]:
            continue
        if sub_categoria and item["sub_categoria"] and sub_categoria != item["sub_categoria"]:
            continue

        mem_norm = item["_origem_memoria_norm"]

        if origem_norm == mem_norm:
            exact.append(item)
            continue

        if strong_partial_match(origem_norm, mem_norm):
            partial.append((token_overlap_score(origem_norm, mem_norm), item))
            continue

        overlap = token_overlap_score(origem_norm, mem_norm)
        if overlap >= 0.60:
            partial.append((overlap, item))

    if exact:
        exact.sort(
            key=lambda x: (
                0 if x["sub_categoria"] == sub_categoria else 1,
                -len(tokenize(x["origem_memoria"])),
            )
        )
        return exact[0]

    if partial:
        partial.sort(
            key=lambda x: (
                -x[0],
                0 if x[1]["sub_categoria"] == sub_categoria else 1,
                -len(tokenize(x[1]["origem_memoria"])),
            )
        )
        return partial[0][1]

    return None


def apply_shadow_mapping(
    rows: list[dict[str, Any]],
    shadow_ws,
) -> list[dict[str, Any]]:
    shadow_items = load_shadow_memory_index(shadow_ws)
    updated = []

    for item in rows:
        row = apply_special_classification_rules(item)

        if str(row.get("destino_template", "")).strip():
            updated.append(row)
            continue

        match = match_shadow_entry(
            origem=str(row.get("origem", "")).strip(),
            grupo=str(row.get("grupo", "")).strip(),
            sub_categoria=str(row.get("sub_categoria", "")).strip(),
            shadow_items=shadow_items,
        )

        if match:
            row["destino_template"] = match["destino_template"]
            row["grupo"] = match["grupo"] or row.get("grupo", "")
            row["sub_categoria"] = match["sub_categoria"] or row.get("sub_categoria", "")
            row["tipo_mapeamento"] = "Memoria Anterior"
            row = apply_special_classification_rules(row)

        updated.append(row)

    return updated


# Linha maxima da Rastreabilidade coberta pelas formulas dinamicas de
# "Memoria Atual" (H/Z) da Shadow. Bound generoso (a aba costuma ter algumas
# centenas de linhas) que ainda cobre acrescimos manuais de linhas feitos pelo
# usuario apos a geracao, sem o custo de varrer a coluna inteira numa formula
# de array. Deve ser >= ao maior numero possivel de linhas da Rastreabilidade.
SHADOW_MEMORIA_MAX_RAST_ROW = 5000


def _memoria_atual_formula(destino_cell: str) -> str:
    """Monta a 'Memoria Atual' DINAMICAMENTE a partir da Rastreabilidade.

    Concatena as Chaves (coluna M = "Origem|Grupo|Sub") de todas as linhas cujo
    "Destino no Template" (K) == destino desta linha da Shadow (A p/ Ativo e
    Passivo, S p/ DRE) e cuja "Alocação da Hierarquia" (D) == "Sim" — a MESMA
    referencia usada pelas colunas de valor B/T do template. Resultado no
    formato "(k1) + (k2) + ...", identico ao texto estatico antigo, de modo que
    a "Memoria Atual Ajustada" (I/AA), que ja consome H/Z, segue valida.

    Escrita como STRING de formula de ARRAY DINAMICO usando _xlfn._xlws.FILTER
    (mesma abordagem da aba Listas, que ja abre sem "@"): o Excel 365 reconhece
    a funcao dinamica e avalia em modo de matriz, SEM injetar a intersecao
    implicita "@". A formula CSE antiga (ArrayFormula, t="array") ainda recebia
    "@" neste ambiente. FILTER sem correspondencia retorna #CALC!, capturado por
    IFERROR -> "" (celula vazia, como o texto estatico antigo)."""
    top = SHADOW_MEMORIA_MAX_RAST_ROW
    return (
        f'=IFERROR("("&_xlfn.TEXTJOIN(") + (",TRUE,'
        f"_xlfn._xlws.FILTER(rastreabilidade!$M$2:$M${top},"
        f"(rastreabilidade!$K$2:$K${top}={destino_cell})"
        f'*(rastreabilidade!$D$2:$D${top}="{ALOCACAO_SIM}")))&")","")'
    )


def update_shadow_memoria_atual(shadow_ws, rows: list[dict[str, Any]]) -> None:
    """Preenche H (Ativo/Passivo) e Z (DRE) com uma FORMULA dinamica que localiza
    as origens alocadas a cada destino na Rastreabilidade, em vez de texto
    estatico colado. Ao realocar/adicionar contas na Rastreabilidade a Memoria
    Atual recalcula sozinha — e, por consequencia, a Memoria Atual Ajustada
    (I/AA, que referencia H/Z). Escreve apenas nas linhas com destino (A/S
    preenchido), como antes.

    A "Memoria Anterior" (C/U) deixa de ser unida aqui; permanece como coluna de
    referencia separada (e segue usada no matching historico de
    apply_shadow_mapping). O parametro `rows` e' mantido por compatibilidade da
    chamada no pipeline."""
    for row in range(5, 40):  # Ativo (destino em A)
        if str(shadow_ws[f"A{row}"].value or "").strip():
            shadow_ws[f"H{row}"] = _memoria_atual_formula(f"$A{row}")

    for row in range(45, 80):  # Passivo (destino em A)
        if str(shadow_ws[f"A{row}"].value or "").strip():
            shadow_ws[f"H{row}"] = _memoria_atual_formula(f"$A{row}")

    for row in range(5, 40):  # DRE (destino em S)
        if str(shadow_ws[f"S{row}"].value or "").strip():
            shadow_ws[f"Z{row}"] = _memoria_atual_formula(f"$S{row}")


# =========================================================
# Rastreabilidade
# =========================================================
def _strip_totalizador_suffix(name: Any) -> str:
    """Remove um sufixo ' - Totalizador' do texto (idempotencia)."""
    return re.sub(
        r"\s*-\s*totalizador\s*$", "", str(name or "").strip(), flags=re.IGNORECASE
    ).strip()


def compute_totalizador_origens(rows: list[dict[str, Any]]) -> set[str]:
    """Origens (normalizadas, sem sufixo) que sao 'pais' (totalizadores):
    aparecem como Hierarquia de pelo menos uma OUTRA linha (abertura). Conta
    de topo sem filhos nao entra (nao e totalizador)."""
    parent_refs: set[str] = set()
    for row in rows:
        hier = normalize_text(_strip_totalizador_suffix(row.get("hierarquia", "")))
        origem = normalize_text(_strip_totalizador_suffix(row.get("origem", "")))
        if hier and hier != origem:
            parent_refs.add(hier)
    return parent_refs


def _hierarquia_display(
    item: dict[str, Any], totalizador_origens: set[str] | None = None
) -> Any:
    """Valor exibido na coluna Hierarquia:
    - TOTALIZADOR (a propria origem e' pai de alguma abertura): nome da PROPRIA
      conta -- INCLUSIVE quando ela tambem e' abertura de um pai de nivel
      superior (multi-nivel). Pedido do especialista: o totalizador encabeca o
      proprio grupo e NAO deve exibir o nome do avo. Ex.: '1.1.1.03 - APLICACOES
      FINANCEIRAS LIQ IMEDIATA' (filho de '1.1.1 - DISPONIVEL', mas pai de outras
      aberturas) fica com Hierarquia = '1.1.1.03 - APLICACOES...', nao
      '1.1.1 - DISPONIVEL'.
    - Abertura (folha): nome da conta-pai imediata.
    - Top-level sem filhos: nome da propria conta.
    O flag de totalizador vai na coluna 'Totalizador' (C). Idempotente (remove
    sufixo ' - Totalizador' preexistente). 'totalizador_origens' vem de
    compute_totalizador_origens (derivado da hierarquia ORIGINAL/pai, intacta
    para os guardrails); quando None, mantem o comportamento antigo (nome do
    pai)."""
    origem = _strip_totalizador_suffix(item.get("origem", ""))
    if totalizador_origens and normalize_text(origem) in totalizador_origens:
        return origem or None
    hier = _strip_totalizador_suffix(item.get("hierarquia", "")) or origem
    return hier or None


def _totalizador_flag(item: dict[str, Any], totalizador_origens: set[str]) -> str:
    """Flag da coluna 'Totalizador' (C): 'Sim' quando a origem e' um pai
    (aparece como Hierarquia de alguma abertura), 'Não' caso contrario."""
    origem = normalize_text(_strip_totalizador_suffix(item.get("origem", "")))
    return ALOCACAO_SIM if origem and origem in totalizador_origens else ALOCACAO_NAO


def append_rastreabilidade(
    ws,
    rows: list[dict[str, Any]],
    years: list[str],
    shadow_ws,
) -> tuple[int, int]:
    """
    Escreve a Rastreabilidade mesclando rows por (origem, grupo,
    sub_categoria, destino_template, tipo_mapeamento) e distribuindo os
    valores nas colunas Ano 1/2/3. Escreve a coluna "Totalizador" (C, derivada
    dos dados) e a "Alocação da Hierarquia" (D, do modelo); ambas ja existem no
    template novo, preservando as colunas referenciadas pelas formulas Shadow/Listas.

    A coluna 'Hierarquia' recebe o nome da conta-pai do OCR quando a
    linha é uma abertura (sub-item). Para linhas TOTALIZADORAS (pais) e
    top-level recebe o nome da PROPRIA conta (nunca vazia) -- inclusive um
    totalizador que tambem seja sub-item de um pai superior (o script força o
    nome proprio, a pedido do especialista).

    Os cabecalhos F1/G1/H1 ("Ano 1"/"Ano 2"/"Ano 3") sao renomeados
    para os anos reais identificados (na ordem cronologica do mais
    antigo para o mais recente). Quando ha menos de 3 anos, os
    placeholders restantes ficam como "Ano N".

    As linhas sao ordenadas conforme o Plano de Contas do template
    (lido da Shadow): Ativo Circulante, Ativo Nao Circulante, Passivo
    Circulante, Passivo Nao Circulante, Passivo PL, DRE. Dentro de
    cada grupo, segue a ordem em que o destino aparece na Shadow;
    linhas sem destino vao para o final do bloco do seu grupo.
    """
    header_map = validate_rastreabilidade_headers(ws)

    # Coluna "Alocação da Hierarquia" (Sim/Não). No template novo ela ja existe
    # na posicao D; este bloco e' apenas FALLBACK (cria no fim) caso um template
    # antigo nao a tenha. Registrada no header_map para escrita por nome.
    if ALOCACAO_HIER_HEADER not in header_map:
        aloc_col = max(header_map.values()) + 1
        header_cell = ws.cell(row=1, column=aloc_col)
        header_cell.value = ALOCACAO_HIER_HEADER
        ref_header = ws.cell(row=1, column=header_map["Chave Destino"])
        if ref_header.has_style:
            header_cell._style = copy(ref_header._style)
        header_map[ALOCACAO_HIER_HEADER] = aloc_col
    aloc_col = header_map[ALOCACAO_HIER_HEADER]

    start_row = first_empty_row(ws, start_row=2, key_col=1)

    # Posicoes das colunas de ano (placeholders "Ano 1"/"Ano 2"/"Ano 3"
    # vem do template; serao renomeadas para os anos reais abaixo).
    year_cols = [header_map["Ano 1"], header_map["Ano 2"], header_map["Ano 3"]]

    # Anos alinhados a DIREITA: o mais recente vai para "Ano 3"; com menos de 3
    # anos, as colunas iniciais (ex.: "Ano 1") preservam o placeholder. As
    # formulas da Shadow casam por VALOR do ano (MATCH em D1:F1), entao o
    # alinhamento a direita e suportado sem alterar formulas.
    year_offset = len(year_cols) - len(years)

    # Renomear cabecalhos das colunas de ano com os anos reais (right-aligned)
    for idx, ano in enumerate(years):
        ws.cell(row=1, column=year_cols[year_offset + idx]).value = ano

    # Mesclar rows por chave estrutural e agrupar valores por ano
    merged = merge_rows_for_rastreabilidade(rows)

    # Ordenar conforme Plano de Contas
    destination_order = build_destination_order(shadow_ws)
    merged = sort_merged_rows(merged, destination_order)

    # Linhas "pai" (totalizadores): a origem aparece como Hierarquia de alguma
    # abertura. Marcadas com "Sim" na coluna "Totalizador" (C) para o
    # especialista distinguir, num relance, o totalizador das aberturas.
    totalizador_origens = compute_totalizador_origens(merged)

    next_row = start_row
    for item in merged:
        item = apply_special_classification_rules(item)
        destino_val = item.get("destino_template") or None
        ws.cell(next_row, header_map["Origem"]).value = item.get("origem")
        # Hierarquia: nome da conta-pai (aberturas); para TOTALIZADOR (pai) e
        # top-level usa o nome da PROPRIA conta (nunca vazia), mesmo em multi-nivel.
        ws.cell(next_row, header_map["Hierarquia"]).value = _hierarquia_display(
            item, totalizador_origens
        )
        ws.cell(next_row, header_map["Totalizador"]).value = _totalizador_flag(
            item, totalizador_origens
        )
        ws.cell(next_row, header_map["Página Referência"]).value = item.get("pagina_referencia")
        ws.cell(next_row, header_map["Grupo"]).value = item.get("grupo")
        ws.cell(next_row, header_map["Sub Categoria"]).value = item.get("sub_categoria")
        ws.cell(next_row, header_map["Destino no Template"]).value = destino_val
        ws.cell(next_row, header_map["Tipo de Mapeamento"]).value = item.get("tipo_mapeamento")
        ws.cell(next_row, aloc_col).value = normalize_alocacao_value(
            item.get("alocacao_hierarquia")
        )

        valores_por_ano = item.get("valores_por_ano", {})
        for idx, ano in enumerate(years):
            cell = ws.cell(next_row, year_cols[year_offset + idx])
            cell.value = coerce_number(valores_por_ano.get(ano))
            cell.number_format = YEAR_VALUE_NUMBER_FORMAT

        next_row += 1

    return start_row, next_row - 1


def propagate_formulas_down(ws, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return

    source_row = 2 if start_row <= 2 else start_row - 1
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        source_value = source_cell.value

        if not (isinstance(source_value, str) and source_value.startswith("=")):
            continue

        for target_row in range(start_row, end_row + 1):
            target_cell = ws.cell(row=target_row, column=col)
            if target_cell.value not in (None, ""):
                continue

            try:
                target_cell.value = Translator(
                    source_value,
                    origin=source_cell.coordinate,
                ).translate_formula(target_cell.coordinate)
            except Exception:
                target_cell.value = source_value

            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)


def _tracking_letters(headers: dict[str, int]) -> dict[str, str]:
    """Letras de coluna usadas nas chaves, derivadas do header_map (robusto a
    reordenacoes do template)."""
    return {
        "origem": get_column_letter(headers["Origem"]),
        "grupo": get_column_letter(headers["Grupo"]),
        "sub": get_column_letter(headers["Sub Categoria"]),
        "destino": get_column_letter(headers["Destino no Template"]),
        "aloc": get_column_letter(headers[ALOCACAO_HIER_HEADER]),
    }


TRACKING_REQUIRED = (
    "Chave",
    "Chave Destino",
    "Origem",
    "Grupo",
    "Sub Categoria",
    "Destino no Template",
    ALOCACAO_HIER_HEADER,
)


def fill_tracking_formulas(ws, start_row: int, end_row: int) -> None:
    headers = build_header_map(ws)

    for col in TRACKING_REQUIRED:
        if col not in headers:
            raise ValueError(f"Coluna '{col}' não encontrada na Rastreabilidade.")

    chave_col = headers["Chave"]
    chave_destino_col = headers["Chave Destino"]
    L = _tracking_letters(headers)

    for row in range(start_row, end_row + 1):
        # Chave/Chave Destino so sao geradas quando Alocação da Hierarquia =
        # "Sim"; nas demais linhas ficam vazias e nao entram nos SUMIFS da
        # Shadow (evita dupla contagem, mesmo com Destino preenchido em todas).
        # Chave = Origem | Grupo | Sub Categoria
        ws.cell(row=row, column=chave_col).value = (
            f'=IF(${L["aloc"]}{row}="{ALOCACAO_SIM}",'
            f'{L["origem"]}{row}&"|"&{L["grupo"]}{row}&"|"&{L["sub"]}{row},"")'
        )
        # Chave Destino = Destino no Template | Grupo | Sub Categoria
        ws.cell(row=row, column=chave_destino_col).value = (
            f'=IF(${L["aloc"]}{row}="{ALOCACAO_SIM}",'
            f'{L["destino"]}{row}&"|"&{L["grupo"]}{row}&"|"&{L["sub"]}{row},"")'
        )


def validate_tracking_columns_position(ws) -> None:
    headers = build_header_map(ws)

    if "Chave" not in headers:
        raise ValueError("Coluna 'Chave' não encontrada na Rastreabilidade.")
    if "Chave Destino" not in headers:
        raise ValueError("Coluna 'Chave Destino' não encontrada na Rastreabilidade.")

    if headers["Chave"] != 13:
        raise ValueError(
            f"A coluna 'Chave' deve estar na posição 13 (M) da Rastreabilidade, mas está em {headers['Chave']}."
        )
    if headers["Chave Destino"] != 14:
        raise ValueError(
            f"A coluna 'Chave Destino' deve estar na posição 14 (N) da Rastreabilidade, mas está em {headers['Chave Destino']}."
        )


# =========================================================
# Anos
# =========================================================
def set_year_headers(shadow_ws, years: list[str]) -> None:
    """
    Escreve os anos reais nos cabecalhos da Shadow (E3:G3 para Ativo/
    Passivo e W3:Y3 para DRE). Recebe a lista de anos ja calculada
    (use compute_years para gerar a partir dos rows do OCR).
    """
    # Right-aligned: o ano mais recente vai para a ULTIMA coluna (G3/Y3); com
    # menos de 3 anos, as colunas iniciais (E3/W3...) ficam vazias. Consistente
    # com a Rastreabilidade; as formulas da Shadow casam por valor do ano.
    ativo_cells = ["E3", "F3", "G3"]
    dre_cells = ["W3", "X3", "Y3"]
    offset = len(ativo_cells) - len(years)
    for i in range(3):
        shadow_ws[ativo_cells[i]] = None
        shadow_ws[dre_cells[i]] = None
    for idx, ano in enumerate(years):
        shadow_ws[ativo_cells[offset + idx]] = ano
        shadow_ws[dre_cells[offset + idx]] = ano


# =========================================================
# Listas
# =========================================================
def apply_listas_formulas(listas_ws) -> None:
    """
    Grava as 3 formulas dinamicas na aba Listas (A2/C2/E2) como ArrayFormula
    com os prefixos _xlfn / _xlfn._xlws / _xlpm exigidos pelo Excel para
    funcoes de array dinamico (LET, SORT, UNIQUE, FILTER, TOCOL, VSTACK).

    A2: lista unica e ordenada das chaves vindas de Rastreabilidade!K:K
    C2: subset de A2 excluindo itens ja usados nas faixas Adicionar da Shadow
    E2: subset de A2 excluindo itens ja usados nas faixas Retirar da Shadow

    Em runtime, as formulas fazem spill nas colunas A/C/E e a Data Validation
    da Shadow consome esses ranges, impedindo selecao em duplicidade.
    """
    f_a2 = (
        '=IFERROR(_xlfn._xlws.SORT(_xlfn.UNIQUE(_xlfn._xlws.FILTER('
        'Rastreabilidade!$M$2:$M$1048576,'
        'Rastreabilidade!$M$2:$M$1048576<>""'
        '))),"")'
    )
    f_c2 = (
        '=IFERROR(_xlfn.LET('
        '_xlpm.Base,_xlfn._xlws.SORT(_xlfn.UNIQUE(_xlfn._xlws.FILTER('
        'Rastreabilidade!$M$2:$M$1048576,'
        'Rastreabilidade!$M$2:$M$1048576<>""))),'
        '_xlpm.Usados,_xlfn.TOCOL(_xlfn.VSTACK('
        'Shadow!$N$5:$Q$39,Shadow!$N$45:$Q$79,Shadow!$AG$5:$AJ$39),1),'
        '_xlfn._xlws.FILTER(_xlpm.Base,ISNA(MATCH(_xlpm.Base,_xlpm.Usados,0)))'
        '),"")'
    )
    f_e2 = (
        '=IFERROR(_xlfn.LET('
        '_xlpm.Base,_xlfn._xlws.SORT(_xlfn.UNIQUE(_xlfn._xlws.FILTER('
        'Rastreabilidade!$M$2:$M$1048576,'
        'Rastreabilidade!$M$2:$M$1048576<>""))),'
        '_xlpm.Usados,_xlfn.TOCOL(_xlfn.VSTACK('
        'Shadow!$J$5:$M$39,Shadow!$J$45:$M$79,Shadow!$AC$5:$AF$39),1),'
        '_xlfn._xlws.FILTER(_xlpm.Base,ISNA(MATCH(_xlpm.Base,_xlpm.Usados,0)))'
        '),"")'
    )

    # Limpar registros pre-existentes de array formula em A2/C2/E2 vindos do
    # template; senao o openpyxl preserva o marcador <f t="array"> e o Excel
    # interpreta como CSE legacy (mostra {=...} na barra e nao faz spill).
    array_formulae = getattr(listas_ws, "array_formulae", None)
    if isinstance(array_formulae, dict):
        for key in ("A2", "C2", "E2"):
            array_formulae.pop(key, None)

    for col_idx in (1, 3, 5):
        for row in range(3, listas_ws.max_row + 1):
            listas_ws.cell(row=row, column=col_idx).value = None

    # Gravar como string plana (NAO ArrayFormula) para que o XML serializado
    # seja <f>FORMULA</f> sem o atributo t="array". Excel 365 detecta as
    # funcoes dinamicas (LET/SORT/UNIQUE/FILTER/TOCOL/VSTACK) e faz spill
    # automatico ao abrir, sem necessidade de F2+Enter.
    listas_ws["A2"] = f_a2
    listas_ws["C2"] = f_c2
    listas_ws["E2"] = f_e2


# =========================================================
# Validação de dados
# =========================================================
def remove_data_validations_in_ranges(shadow_ws, target_ranges: list[str]) -> None:
    dvs = getattr(shadow_ws.data_validations, "dataValidation", [])
    if not dvs:
        return

    remaining = []
    for dv in list(dvs):
        sqref = str(getattr(dv, "sqref", "") or "")
        should_remove = any(target in sqref for target in target_ranges)
        if not should_remove:
            remaining.append(dv)

    shadow_ws.data_validations.dataValidation = remaining


def apply_shadow_data_validations(shadow_ws) -> None:
    retirar_formula = "listas!$E$2:$E$1048576"
    adicionar_formula = "listas!$C$2:$C$1048576"

    retirar_ranges = ["J5:M39", "J45:M79", "AC5:AF39"]
    adicionar_ranges = ["N5:Q39", "N45:Q79", "AG5:AJ39"]

    remove_data_validations_in_ranges(shadow_ws, retirar_ranges + adicionar_ranges)

    for cell_range in retirar_ranges:
        dv = DataValidation(
            type="list",
            formula1=retirar_formula,
            allow_blank=True,
            showDropDown=False,
        )
        shadow_ws.add_data_validation(dv)
        dv.add(cell_range)

    for cell_range in adicionar_ranges:
        dv = DataValidation(
            type="list",
            formula1=adicionar_formula,
            allow_blank=True,
            showDropDown=False,
        )
        shadow_ws.add_data_validation(dv)
        dv.add(cell_range)


def verify_shadow_validations_config(shadow_ws) -> None:
    expected = {
        "J5:M39": "listas!$E$2:$E$1048576",
        "J45:M79": "listas!$E$2:$E$1048576",
        "AC5:AF39": "listas!$E$2:$E$1048576",
        "N5:Q39": "listas!$C$2:$C$1048576",
        "N45:Q79": "listas!$C$2:$C$1048576",
        "AG5:AJ39": "listas!$C$2:$C$1048576",
    }

    existing = []
    for dv in getattr(shadow_ws.data_validations, "dataValidation", []):
        existing.append((str(getattr(dv, "sqref", "") or ""), str(getattr(dv, "formula1", "") or "").lstrip("=")))

    for target_range, target_formula in expected.items():
        found = any(target_range in sqref and formula1 == target_formula for sqref, formula1 in existing)
        if not found:
            raise RuntimeError(
                f"Validação obrigatória não encontrada para {target_range} com fórmula {target_formula}"
            )


# =========================================================
# Integridade
# =========================================================
def snapshot_protected_ranges(ws) -> dict[str, Any]:
    protected = {}
    ranges = [("I", "Q"), ("AA", "AJ")]
    max_row = ws.max_row

    for col_start, col_end in ranges:
        c1 = column_index_from_string(col_start)
        c2 = column_index_from_string(col_end)
        for row in range(1, max_row + 1):
            for col in range(c1, c2 + 1):
                coord = ws.cell(row=row, column=col).coordinate
                protected[coord] = ws.cell(row=row, column=col).value

    return protected


SHEET_NAME_NORMALIZATIONS = {
    "shadow": "Shadow",
    "rastreabilidade": "Rastreabilidade",
    "listas": "Listas",
}


def normalize_sheet_names(wb) -> None:
    """
    Capitaliza a primeira letra dos nomes das sheets canonicas para que o
    Excel exiba referencias como `Rastreabilidade!`, `Shadow!`, `Listas!` ao
    inves de `rastreabilidade!`, `shadow!`, `listas!` na barra de formulas.

    Necessario porque o Excel sempre re-renderiza referencias de formula
    conforme o nome real da sheet (case-preserving), ignorando o case com que
    a referencia foi escrita na string da formula. Como as formulas em
    `Listas!A2/C2/E2` ja sao gravadas com nomes capitalizados, basta renomear
    as proprias sheets para que o display case na abertura no Excel coincida.

    Idempotente: se ja capitalizado, nao altera. Seguro de chamar antes de
    `wb.save`. Resolucao de referencias por sheet name e case-insensitive em
    Excel, entao formulas pre-existentes em outras sheets continuam funcionando.

    Implementacao em 2 etapas: openpyxl considera nomes case-insensitive ao
    detectar colisoes (ex.: tentar renomear `shadow` para `Shadow` faz a
    biblioteca achar que ha colisao com a propria sheet sendo renomeada e
    adicionar sufixo `1` -> `Shadow1`). Para contornar, primeiro renomeamos
    para um nome unico temporario e depois para o final.
    """
    pending: dict[str, str] = {}
    for ws in wb.worksheets:
        target = SHEET_NAME_NORMALIZATIONS.get(ws.title.lower())
        if target and ws.title != target:
            tmp = f"__norm_tmp__{ws.title}__"
            ws.title = tmp
            pending[tmp] = target

    for tmp, final in pending.items():
        wb[tmp].title = final


DA_METADATA_XML = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<metadata xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:xda="http://schemas.microsoft.com/office/spreadsheetml/2017/dynamicarray">
  <metadataTypes count="1">
    <metadataType name="XLDAPR" minSupportedVersion="120000" copy="1" pasteAll="1" pasteValues="1" merge="1" splitFirst="1" rowColShift="1" clearFormats="1" clearComments="1" assign="1" coerce="1" cellMeta="1"/>
  </metadataTypes>
  <futureMetadata name="XLDAPR" count="1">
    <bk>
      <extLst>
        <ext uri="{bdbb8cdc-fa1e-496e-a857-3c3f30c029c3}">
          <xda:dynamicArrayProperties fDynamic="1" fCollapsed="0"/>
        </ext>
      </extLst>
    </bk>
  </futureMetadata>
  <cellMetadata count="1">
    <bk>
      <rc t="1" v="0"/>
    </bk>
  </cellMetadata>
</metadata>"""


def apply_dynamic_array_metadata(xlsx_path: Path) -> None:
    """
    Pos-processa o xlsx adicionando metadados OOXML de Dynamic Array nas
    formulas de Listas!A2/C2/E2.

    Sem esses metadados, o Excel ao abrir adiciona o operador de intersecao
    implicita (`@`) antes de funcoes que retornam array (LET, SORT, UNIQUE,
    FILTER, TOCOL, VSTACK), forcando avaliacao escalar e impedindo o spill.
    O usuario precisaria editar cada celula manualmente (F2 -> remover @ ->
    Enter) para que a formula fizesse spill.

    Esta funcao adiciona:

    - `xl/metadata.xml` com metadataType `XLDAPR` + `dynamicArrayProperties
      fDynamic="1"`
    - `<Override>` em `[Content_Types].xml` apontando para metadata.xml
    - `<Relationship>` em `xl/_rels/workbook.xml.rels` apontando para
      metadata.xml com o type sheetMetadata
    - atributo `cm="1"` em cada celula A2/C2/E2 da aba listas, ligando
      essas celulas ao bloco cellMetadata do metadata.xml

    Idempotente: se ja existir, nao duplica.
    """
    xlsx_path = Path(xlsx_path)
    tmp = xlsx_path.with_suffix(".tmp.xlsx")

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = list(zin.namelist())
        files = {n: zin.read(n) for n in names}

    # 1. Identificar a sheet Listas por conteudo (mais robusto que parsear
    #    workbook.xml + rels). A aba Listas contem em A2/C2/E2 formulas com
    #    FILTER + prefixos _xlfn.
    sheet_path = None
    for n, data in files.items():
        if not (n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
            continue
        s = data.decode("utf-8", errors="replace")
        a2 = re.search(r'<c r="A2"[^>]*>.*?</c>', s, re.DOTALL)
        c2 = re.search(r'<c r="C2"[^>]*>.*?</c>', s, re.DOTALL)
        e2 = re.search(r'<c r="E2"[^>]*>.*?</c>', s, re.DOTALL)
        if a2 and c2 and e2:
            joined = (a2.group() + c2.group() + e2.group())
            if "FILTER" in joined.upper() and "_xlfn" in joined:
                sheet_path = n
                break

    if sheet_path is None:
        # Aba Listas nao tem as formulas (ex.: template incompleto). Nada a fazer.
        return

    # 2. xl/metadata.xml (sobrescreve se ja existir, garantindo conteudo conhecido)
    files["xl/metadata.xml"] = DA_METADATA_XML

    # 3. [Content_Types].xml - Override para metadata.xml
    ct = files["[Content_Types].xml"].decode("utf-8")
    if "/xl/metadata.xml" not in ct:
        ovr = (
            '<Override PartName="/xl/metadata.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheetMetadata+xml"/>'
        )
        ct = ct.replace("</Types>", ovr + "</Types>")
        files["[Content_Types].xml"] = ct.encode("utf-8")

    # 4. xl/_rels/workbook.xml.rels - Relationship para metadata.xml
    rels_path = "xl/_rels/workbook.xml.rels"
    rels = files[rels_path].decode("utf-8")
    if "metadata.xml" not in rels:
        used_ids = [int(i) for i in re.findall(r'Id="rId(\d+)"', rels)]
        next_id = (max(used_ids) + 1) if used_ids else 1
        rel = (
            f'<Relationship Id="rId{next_id}" '
            f'Type="http://schemas.microsoft.com/office/2017/06/relationships/sheetMetadata" '
            f'Target="metadata.xml"/>'
        )
        rels = rels.replace("</Relationships>", rel + "</Relationships>")
        files[rels_path] = rels.encode("utf-8")

    # 5. cm="1" nas celulas A2/C2/E2 da Listas
    sheet = files[sheet_path].decode("utf-8")
    for ref in ("A2", "C2", "E2"):
        if f'<c r="{ref}" cm="1"' in sheet:
            continue
        pat = rf'<c r="{ref}"((?:(?!cm=")[^>])*?)>'
        sheet = re.sub(pat, rf'<c r="{ref}" cm="1"\1>', sheet, count=1)
    files[sheet_path] = sheet.encode("utf-8")

    # 6. Reescrever o zip preservando ordem dos arquivos
    final_names = list(names)
    if "xl/metadata.xml" not in names:
        # inserir apos workbook.xml para ficar perto dos outros artefatos do xl/
        try:
            idx = final_names.index("xl/workbook.xml") + 1
            final_names.insert(idx, "xl/metadata.xml")
        except ValueError:
            final_names.append("xl/metadata.xml")

    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in final_names:
            zout.writestr(n, files[n])

    shutil.move(str(tmp), str(xlsx_path))


WORKBOOK_CALC_FEATURES_EXT = (
    '<extLst>'
    '<ext uri="{B58B0392-4F1F-4190-BB64-5DF3571DCE5F}" '
    'xmlns:xcalcf="http://schemas.microsoft.com/office/spreadsheetml/2018/calcfeatures">'
    '<xcalcf:calcFeatures>'
    '<xcalcf:feature name="microsoft.com:RD"/>'
    '<xcalcf:feature name="microsoft.com:Single"/>'
    '<xcalcf:feature name="microsoft.com:FV"/>'
    '<xcalcf:feature name="microsoft.com:CNMTM"/>'
    '<xcalcf:feature name="microsoft.com:LET_WF"/>'
    '<xcalcf:feature name="microsoft.com:LAMBDA_WF"/>'
    '<xcalcf:feature name="microsoft.com:ARRAYTEXT_WF"/>'
    '</xcalcf:calcFeatures>'
    '</ext>'
    '</extLst>'
)


def apply_workbook_calc_features(xlsx_path: Path) -> None:
    """
    Adiciona o <extLst> com calcFeatures em xl/workbook.xml para que o
    Excel 365 reconheca o arquivo como capaz de Dynamic Arrays e nao
    insira o operador de intersecao implicita ('@') antes das funcoes
    LET / SORT / UNIQUE / FILTER / TOCOL / VSTACK em Listas!A2/C2/E2.

    A feature critica e 'microsoft.com:RD' (Reduce/Dynamic). As demais
    sao declaracoes defensivas para builds que verificam features
    especificas (LET_WF, LAMBDA_WF, ARRAYTEXT_WF, FV, CNMTM, Single).

    Sem este extLst, mesmo com xl/metadata.xml + cm="1" + Override em
    [Content_Types].xml + Relationship em workbook.xml.rels todos
    corretos, o Excel ainda trata o workbook em modo legacy e adiciona
    o '@' ao abrir, travando o spill ate que o usuario edite cada
    celula manualmente.

    Idempotente: se calcFeatures ja existir, nao duplica.
    """
    xlsx_path = Path(xlsx_path)
    tmp = xlsx_path.with_suffix(".tmp.xlsx")

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = list(zin.namelist())
        files = {n: zin.read(n) for n in names}

    wb_xml = files["xl/workbook.xml"].decode("utf-8")

    if "calcFeatures" in wb_xml:
        return

    if "</workbook>" not in wb_xml:
        return

    wb_xml = wb_xml.replace(
        "</workbook>", WORKBOOK_CALC_FEATURES_EXT + "</workbook>"
    )
    files["xl/workbook.xml"] = wb_xml.encode("utf-8")

    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, files[n])

    shutil.move(str(tmp), str(xlsx_path))


def _read_template_listas_cells(template_path: Path | None) -> dict[str, str]:
    """
    Le, do TEMPLATE, o XML verbatim das celulas A2/C2/E2 da aba Listas (formula
    dinamica + atributo cm="1", exatamente como o usuario gravou). Serve para
    reimpor esse conteudo no arquivo final, garantindo que a formula gerada seja
    IDENTICA a do template (o script nunca reescreve/altera a formula). Retorna {}
    se o template nao for legivel ou nao tiver a aba Listas com formulas dinamicas
    (FILTER + _xlfn em A2/C2/E2).
    """
    if not template_path:
        return {}
    try:
        with zipfile.ZipFile(Path(template_path), "r") as z:
            for n in z.namelist():
                if not (n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
                    continue
                s = z.read(n).decode("utf-8", errors="replace")
                cells: dict[str, str] = {}
                for ref in ("A2", "C2", "E2"):
                    m = re.search(rf'<c r="{ref}"[^>]*>.*?</c>', s, re.DOTALL)
                    if m is None:
                        m = re.search(rf'<c r="{ref}"[^>]*/>', s)
                    if m:
                        cells[ref] = m.group()
                joined = "".join(cells.values())
                if cells.get("A2") and "FILTER" in joined.upper() and "_xlfn" in joined:
                    return cells
    except Exception:
        pass
    return {}


def _replace_cell_xml(sheet_xml: str, ref: str, new_cell: str) -> str:
    """
    Substitui no XML da planilha a celula <c r="ref" ...>...</c> (ou auto-fechada
    <c r="ref" .../>) pelo conteudo `new_cell`. Usa funcao de replacement para nao
    interpretar barras/grupos presentes na formula. Retorna o XML inalterado se a
    celula nao for encontrada.
    """
    pat_full = rf'<c r="{ref}"[^>]*>.*?</c>'
    if re.search(pat_full, sheet_xml, re.DOTALL):
        return re.sub(pat_full, lambda _m: new_cell, sheet_xml, count=1, flags=re.DOTALL)
    pat_self = rf'<c r="{ref}"[^>]*/>'
    if re.search(pat_self, sheet_xml):
        return re.sub(pat_self, lambda _m: new_cell, sheet_xml, count=1)
    return sheet_xml


def apply_dynamic_array_artifacts(xlsx_path: Path, template_path: Path | None = None) -> None:
    """
    Garante que as celulas A2/C2/E2 da aba Listas no arquivo final fiquem
    IDENTICAS as do template, SEM adicionar nenhum artefato de Dynamic Array.

    Por que nao adicionar nada: as formulas da Listas usam funcoes
    _xlfn._xlws.* (SORT/FILTER/UNIQUE/LET/TOCOL/VSTACK), que o Excel 365 ja
    reconhece como Dynamic Array pelo proprio nome e faz spill sem inserir o
    operador de intersecao implicita "@". O proprio template do usuario NAO tem
    xl/metadata.xml, NEM cm nas celulas, NEM calcFeatures, e abre corretamente.

    Versoes anteriores adicionavam esses 3 artefatos (cm="1" nas celulas +
    xl/metadata.xml + calcFeatures). Isso fazia o Excel REPARAR a planilha ao
    abrir ("Repaired Records: Cell information from /xl/worksheets/sheet3.xml")
    porque o cm referencia um registro de metadata.xml e o Excel valida/rejeita
    essa combinacao. Portanto este passo APENAS reimpoe A2/C2/E2 verbatim do
    template (desfazendo qualquer alteracao do round-trip do openpyxl) e nao
    escreve mais nada — nunca toca em metadata.xml, Content_Types, rels ou
    workbook.xml.

    Idempotente e seguro: se o template nao for legivel ou a aba Listas nao for
    localizada, nao faz nada.
    """
    if template_path is None:
        return
    template_cells = _read_template_listas_cells(template_path)
    if not template_cells:
        return

    xlsx_path = Path(xlsx_path)
    tmp = xlsx_path.with_suffix(".tmp.xlsx")

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = list(zin.namelist())
        files = {n: zin.read(n) for n in names}

    # Localizar a sheet Listas por conteudo (FILTER + _xlfn em A2/C2/E2).
    sheet_path = None
    for n, data in files.items():
        if not (n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
            continue
        s = data.decode("utf-8", errors="replace")
        a2 = re.search(r'<c r="A2"[^>]*>.*?</c>', s, re.DOTALL)
        c2 = re.search(r'<c r="C2"[^>]*>.*?</c>', s, re.DOTALL)
        e2 = re.search(r'<c r="E2"[^>]*>.*?</c>', s, re.DOTALL)
        if a2 and c2 and e2:
            joined = a2.group() + c2.group() + e2.group()
            if "FILTER" in joined.upper() and "_xlfn" in joined:
                sheet_path = n
                break

    if sheet_path is None:
        return

    # Reimpor A2/C2/E2 verbatim do template (formula exatamente como o usuario
    # gravou, SEM cm). Nenhum outro arquivo do zip e' tocado.
    sheet = files[sheet_path].decode("utf-8")
    new_sheet = sheet
    for _ref, _cell in template_cells.items():
        new_sheet = _replace_cell_xml(new_sheet, _ref, _cell)

    if new_sheet == sheet:
        return  # ja identico ao template: nada a reescrever (poupa IO + sync)

    files[sheet_path] = new_sheet.encode("utf-8")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, files[n])

    shutil.move(str(tmp), str(xlsx_path))


def verify_dynamic_array_artifacts(xlsx_path: Path, template_path: Path | None = None) -> None:
    """
    Pos-condicao da aba Listas: confirma que A2/C2/E2 estao IDENTICAS as do
    template e que NENHUMA celula recebeu metadado que faria o Excel reparar o
    arquivo.

    As formulas _xlfn._xlws.* (SORT/FILTER/UNIQUE/LET/...) ja fazem spill
    sozinhas e o template do usuario nao tem cm/metadata.xml/calcFeatures. Por
    isso, em vez de EXIGIR esses artefatos (o que versoes antigas faziam e
    CORROMPIA o sheet3), aqui PROIBIMOS:
      (a) cm em qualquer celula sem formula (<f>) -> "Repaired Records";
      (b) cm nas celulas A2/C2/E2 da Listas -> foi a causa da corrupcao do sheet3;
      (c) operador "@" / _xlfn.SINGLE em qualquer formula;
    e EXIGIMOS que A2/C2/E2 sejam identicas as do template.

    Se a aba Listas nao for encontrada, retorna silenciosamente. Qualquer
    problema gera RuntimeError explicito.
    """
    xlsx_path = Path(xlsx_path)
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        files = {n: zin.read(n) for n in zin.namelist()}

    sheet = None
    for n, data in files.items():
        if not (n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
            continue
        s = data.decode("utf-8", errors="replace")
        a2 = re.search(r'<c r="A2"[^>]*>.*?</c>', s, re.DOTALL)
        c2 = re.search(r'<c r="C2"[^>]*>.*?</c>', s, re.DOTALL)
        e2 = re.search(r'<c r="E2"[^>]*>.*?</c>', s, re.DOTALL)
        if a2 and c2 and e2:
            joined = a2.group() + c2.group() + e2.group()
            if "FILTER" in joined.upper() and "_xlfn" in joined:
                sheet = s
                break

    if sheet is None:
        return

    problems = []
    # cm nas celulas A2/C2/E2 da Listas FOI a causa do "Repaired Records" no
    # sheet3: o Excel valida o cm contra xl/metadata.xml e repara a planilha. As
    # formulas _xlfn._xlws.* ja fazem spill sem cm/metadata (o template do
    # usuario nao tem nenhum desses artefatos), entao aqui PROIBIMOS o cm.
    for ref in ("A2", "C2", "E2"):
        _cm = re.search(rf'<c r="{ref}"[^>]*?>', sheet)
        if _cm and "cm=" in _cm.group():
            problems.append(f'cm indevido em Listas!{ref} (corromperia o sheet3)')

    # Garantias extras (varredura via parser XML, nao regex, para nao recair em
    # bugs com celulas auto-fechadas <c .../>):
    #  (a) ANTI-CORRUPCAO: nenhuma celula pode ter cm="..." sem uma formula (<f>).
    #      O cm referencia metadado de Dynamic Array; numa celula vazia o Excel
    #      considera invalido e "repara" o arquivo (dialogo "Repaired Records:
    #      Cell information", perda de dados).
    #  (b) SEM "@": nenhuma formula pode conter o operador de intersecao implicita
    #      "@" (nem _xlfn.SINGLE). Nosso pipeline nunca o grava, mas se um template
    #      for reaberto/salvo pelo Excel com "@" ele persistiria no texto da
    #      formula. Falha explicita (nunca faz replace cego de "@").
    _ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    for _n, _data in files.items():
        if not (_n.startswith("xl/worksheets/sheet") and _n.endswith(".xml")):
            continue
        try:
            _root = ET.fromstring(_data)
        except ET.ParseError:
            continue
        _cm_sem_f, _com_arroba = [], []
        for _c in _root.iter(_ns + "c"):
            _f = _c.find(_ns + "f")
            if _c.get("cm") is not None and _f is None:
                _cm_sem_f.append(_c.get("r"))
            if _f is not None and _f.text and (
                "@" in _f.text or "_xlfn.SINGLE" in _f.text
            ):
                _com_arroba.append(_c.get("r"))
        if _cm_sem_f:
            problems.append(
                f"cm em celula sem formula em {_n} (Excel repararia/corromperia "
                f"o arquivo): {_cm_sem_f[:15]}"
            )
        if _com_arroba:
            problems.append(
                f"operador de intersecao implicita '@' presente em formulas de "
                f"{_n}: {_com_arroba[:15]}"
            )

    # Garantia extra: a formula de A2/C2/E2 no output deve ser identica a do
    # template (o script nao deve alterar as formulas da Listas).
    if template_path is not None:
        for _ref, _tpl in _read_template_listas_cells(template_path).items():
            _tplf = re.search(r"<f[^>]*>.*?</f>", _tpl, re.DOTALL)
            _outc = re.search(rf'<c r="{_ref}"[^>]*>.*?</c>', sheet, re.DOTALL)
            _outf = re.search(r"<f[^>]*>.*?</f>", _outc.group(), re.DOTALL) if _outc else None
            if _tplf and (_outf is None or _outf.group() != _tplf.group()):
                problems.append(f"formula de Listas!{_ref} difere do template")

    if problems:
        raise RuntimeError(
            "Verificacao da aba Listas falhou (o arquivo final seria reparado "
            "pelo Excel ou abriria com '@'): " + "; ".join(problems)
        )


def verify_shadow_integrity(before_protected: dict[str, Any], output_path: Path, shadow_sheet_name: str) -> None:
    wb_check = open_workbook(output_path)
    shadow_check = wb_check[shadow_sheet_name]
    listas_check = get_sheet_by_name_case_insensitive(wb_check, "listas")

    verify_shadow_validations_config(shadow_check)

    allowed_validation_columns = {
        "J", "K", "L", "M",
        "N", "O", "P", "Q",
        "AC", "AD", "AE", "AF",
        "AG", "AH", "AI", "AJ",
    }

    for coord, before_value in before_protected.items():
        col_letters = "".join(c for c in coord if c.isalpha())
        after_value = shadow_check[coord].value

        if col_letters in allowed_validation_columns:
            if before_value != after_value:
                raise RuntimeError(f"Valor alterado indevidamente na Shadow: {coord}")
            continue

        if before_value != after_value:
            raise RuntimeError(f"Faixa protegida alterada indevidamente na Shadow: {coord}")

    for cell in ("A2", "C2", "E2"):
        if listas_check[cell].value in (None, ""):
            continue


# =========================================================
# QA
# =========================================================
def validate_subcategoria(rows: list[dict[str, Any]]) -> list[str]:
    issues = []
    for idx, row in enumerate(rows, start=1):
        row = apply_special_classification_rules(row)
        grupo = str(row.get("grupo", "")).strip()
        sub = str(row.get("sub_categoria", "")).strip()

        if grupo == "Ativo" and sub not in {"Circulante", "Não Circulante"}:
            issues.append(f"Linha {idx}: Grupo {grupo} sem Sub Categoria válida.")
        if grupo == "Passivo" and sub not in {"Circulante", "Não Circulante", "PL"}:
            issues.append(f"Linha {idx}: Grupo {grupo} sem Sub Categoria válida.")
        if grupo == "DRE" and sub != "DRE":
            issues.append(f"Linha {idx}: Grupo DRE deve usar Sub Categoria = DRE.")
    return issues


def validate_duplicate_allocation(rows: list[dict[str, Any]]) -> list[str]:
    seen = {}
    issues = []

    for idx, row in enumerate(rows, start=1):
        if not row.get("destino_template"):
            continue

        key = (
            normalize_text(row.get("origem", "")),
            str(row.get("pagina_referencia", "")).strip(),
            str(row.get("valor", "")).strip(),
            str(row.get("ano", "")).strip(),
        )
        destino = str(row.get("destino_template", "")).strip()
        grupo = str(row.get("grupo", "")).strip()
        prev = seen.get(key)

        if prev and prev != (destino, grupo):
            issues.append(f"Linha {idx}: possível dupla alocação/conflito para origem {row.get('origem')}.")
        else:
            seen[key] = (destino, grupo)

    return issues


def validate_group_swaps(rows: list[dict[str, Any]], dictionary_rows: list[dict[str, str]]) -> list[str]:
    issues = []
    dict_group_map = defaultdict(set)

    for row in dictionary_rows:
        dict_group_map[row["_origem_norm"]].add(row["Grupo"])

    for idx, row in enumerate(rows, start=1):
        origem_norm = normalize_text(row.get("origem", ""))
        grupo = str(row.get("grupo", "")).strip()

        if origem_norm in dict_group_map and grupo and grupo not in dict_group_map[origem_norm]:
            issues.append(
                f"Linha {idx}: grupo incompatível com o dicionário para origem {row.get('origem')}."
            )

    return issues


def validate_template_destination_structure(rows: list[dict[str, Any]], template_index) -> list[str]:
    issues = []
    for idx, row in enumerate(rows, start=1):
        row = apply_special_classification_rules(row)
        destino = str(row.get("destino_template", "")).strip()
        if not destino:
            continue
        ok = is_destination_compatible(
            destino=destino,
            grupo=row.get("grupo", ""),
            sub_categoria=row.get("sub_categoria", ""),
            template_index=template_index,
        )
        if not ok:
            issues.append(
                f"Linha {idx}: destino '{destino}' incompatível com Grupo/Sub Categoria da origem."
            )
    return issues


def validate_group_is_structurally_correct(rows, template_index) -> list[str]:
    issues = []
    for idx, row in enumerate(rows, start=1):
        row = apply_special_classification_rules(row)
        destino = str(row.get("destino_template", "")).strip()
        if not destino:
            continue
        if not is_destination_compatible(
            destino,
            row.get("grupo", ""),
            row.get("sub_categoria", ""),
            template_index,
        ):
            issues.append(
                f"Linha {idx}: alocação estrutural inválida para destino {destino}."
            )
    return issues


def validate_tracking_formulas(ws, start_row: int, end_row: int) -> list[str]:
    issues = []
    headers = build_header_map(ws)

    if not set(TRACKING_REQUIRED).issubset(headers):
        return issues

    chave_col = headers["Chave"]
    chave_destino_col = headers["Chave Destino"]
    L = _tracking_letters(headers)

    for row in range(start_row, end_row + 1):
        # Chave = IF(Alocação="Sim"; Origem | Grupo | Sub Categoria; "")
        expected_chave = (
            f'=IF(${L["aloc"]}{row}="{ALOCACAO_SIM}",'
            f'{L["origem"]}{row}&"|"&{L["grupo"]}{row}&"|"&{L["sub"]}{row},"")'
        )
        if ws.cell(row=row, column=chave_col).value != expected_chave:
            issues.append(f"Rastreabilidade linha {row}: fórmula da coluna Chave incorreta.")

        # Chave Destino = IF(Alocação="Sim"; Destino | Grupo | Sub Categoria; "")
        expected_chave_destino = (
            f'=IF(${L["aloc"]}{row}="{ALOCACAO_SIM}",'
            f'{L["destino"]}{row}&"|"&{L["grupo"]}{row}&"|"&{L["sub"]}{row},"")'
        )
        if ws.cell(row=row, column=chave_destino_col).value != expected_chave_destino:
            issues.append(f"Rastreabilidade linha {row}: fórmula da coluna Chave Destino incorreta.")

    return issues


def validate_shadow_memory_structure(shadow_ws) -> list[str]:
    issues = []

    for row in range(5, 40):
        memoria = split_memory(shadow_ws[f"H{row}"].value)
        expected_group = "Ativo"
        expected_sub = infer_subcategoria_from_position("Ativo", row)

        for item in memoria:
            parts = [p.strip() for p in item.split("|")]
            if len(parts) >= 3:
                _, grupo, sub = parts[0], parts[1], parts[2]
                if grupo != expected_group:
                    issues.append(
                        f"Shadow H{row}: memória com grupo incompatível. Esperado {expected_group}, encontrado {grupo}."
                    )
                if sub != expected_sub:
                    issues.append(
                        f"Shadow H{row}: memória com subcategoria incompatível. Esperado {expected_sub}, encontrado {sub}."
                    )

    for row in range(45, 80):
        memoria = split_memory(shadow_ws[f"H{row}"].value)
        expected_group = "Passivo"
        expected_sub = infer_subcategoria_from_position("Passivo", row)

        for item in memoria:
            parts = [p.strip() for p in item.split("|")]
            if len(parts) >= 3:
                _, grupo, sub = parts[0], parts[1], parts[2]
                if grupo != expected_group:
                    issues.append(
                        f"Shadow H{row}: memória com grupo incompatível. Esperado {expected_group}, encontrado {grupo}."
                    )
                if sub != expected_sub:
                    issues.append(
                        f"Shadow H{row}: memória com subcategoria incompatível. Esperado {expected_sub}, encontrado {sub}."
                    )

    for row in range(5, 40):
        memoria = split_memory(shadow_ws[f"Z{row}"].value)
        expected_group = "DRE"
        expected_sub = "DRE"

        for item in memoria:
            parts = [p.strip() for p in item.split("|")]
            if len(parts) >= 3:
                _, grupo, sub = parts[0], parts[1], parts[2]
                if grupo != expected_group:
                    issues.append(
                        f"Shadow Z{row}: memória com grupo incompatível. Esperado {expected_group}, encontrado {grupo}."
                    )
                if sub != expected_sub:
                    issues.append(
                        f"Shadow Z{row}: memória com subcategoria incompatível. Esperado {expected_sub}, encontrado {sub}."
                    )

    return issues


def validate_listas_formulas(listas_ws) -> list[str]:
    issues = []
    for cell in ("A2", "C2", "E2"):
        value = listas_ws[cell].value
        if isinstance(value, ArrayFormula):
            text = (value.text or "").upper()
        elif isinstance(value, str):
            text = value.upper()
        else:
            text = ""
        if not text.startswith("="):
            issues.append(
                f"Listas!{cell}: esperado formula dinamica, encontrado {type(value).__name__}."
            )
        elif "FILTER" not in text:
            issues.append(
                f"Listas!{cell}: formula nao parece ser de array dinamico (sem FILTER)."
            )
    return issues


def validate_pl_specific_accounts(rows: list[dict[str, Any]]) -> list[str]:
    issues = []
    for idx, row in enumerate(rows, start=1):
        origem = str(row.get("origem", "")).strip()
        if is_pl_specific_account(origem):
            grupo = str(row.get("grupo", "")).strip()
            sub = str(row.get("sub_categoria", "")).strip()
            if grupo != "Passivo" or sub != "PL":
                issues.append(
                    f"Linha {idx}: conta de PL específica '{origem}' deve ficar com Grupo=Passivo e Sub Categoria=PL."
                )
    return issues


def validate_dre_subcategory(rows: list[dict[str, Any]]) -> list[str]:
    issues = []
    for idx, row in enumerate(rows, start=1):
        grupo = str(row.get("grupo", "")).strip()
        sub = str(row.get("sub_categoria", "")).strip()
        if grupo == "DRE" and sub != "DRE":
            issues.append(f"Linha {idx}: conta DRE sem Sub Categoria = DRE.")
    return issues


def validate_sibling_consistency(rows: list[dict[str, Any]]) -> list[str]:
    """Guardrail read-only: avisa quando aberturas com a MESMA conta-pai
    (Hierarquia) e mesmo Grupo foram alocadas a destinos DIFERENTES, forte
    indicio de classificacao que ignora a hierarquia do documento (ex.: filhos
    de 'Despesas Gerais e Administrativas' indo parte para '- Despesas
    Administrativas' e parte para '- Despesas com Vendas'). So atua quando a
    coluna Hierarquia esta preenchida."""
    grupos: dict[tuple[str, str], set[str]] = defaultdict(set)
    exemplos: dict[tuple[str, str], list[str]] = defaultdict(list)
    rotulo: dict[tuple[str, str], str] = {}
    for row in rows:
        hier = str(row.get("hierarquia", "")).strip()
        destino = str(row.get("destino_template", "")).strip()
        if not hier or not destino:
            continue
        grupo = str(row.get("grupo", "")).strip()
        key = (normalize_text(hier), grupo)
        grupos[key].add(normalize_text(destino))
        exemplos[key].append(f"{str(row.get('origem', '')).strip()} -> {destino}")
        rotulo.setdefault(key, hier)
    issues = []
    for key, destinos in grupos.items():
        if len(destinos) > 1:
            issues.append(
                f"Hierarquia '{rotulo[key]}' (Grupo {key[1] or '?'}): aberturas "
                f"alocadas a destinos diferentes [{'; '.join(exemplos[key])}]. "
                f"Verificar se devem seguir a classificacao da conta-pai."
            )
    return issues


def validate_gpt_vs_dictionary(
    rows: list[dict[str, Any]],
    dictionary_rows: list[dict[str, str]],
) -> list[str]:
    """Guardrail read-only: sinaliza quando uma linha foi alocada num destino
    que DIVERGE do que o dicionario sugere para a mesma origem (mesmo
    Grupo/Sub Categoria). Nao altera nada; serve para revisao do analista.
    Linhas 'Memoria Anterior' sao ignoradas (tem prioridade sobre o dicionario)."""
    issues = []
    for idx, row in enumerate(rows, start=1):
        destino = str(row.get("destino_template", "")).strip()
        if not destino:
            continue
        if str(row.get("tipo_mapeamento", "")).strip() == "Memoria Anterior":
            continue
        match = match_dictionary_entry(
            origem=str(row.get("origem", "")).strip(),
            grupo=str(row.get("grupo", "")).strip(),
            sub_categoria=str(row.get("sub_categoria", "")).strip(),
            dictionary_rows=dictionary_rows,
        )
        if match and normalize_text(match["Destino no Template"]) != normalize_text(destino):
            issues.append(
                f"Linha {idx}: origem '{row.get('origem')}' alocada em "
                f"'{destino}', mas o dicionario sugere "
                f"'{match['Destino no Template']}'. Revisar coerencia."
            )
    return issues


def validate_alocacao_consistency(rows: list[dict[str, Any]]) -> list[str]:
    """Coerencia da 'Alocação da Hierarquia' (read-only, vai p/ o parecer):
    anti-dupla-contagem -- dentro da mesma Hierarquia/Grupo, totalizador
    (Origem == Hierarquia das aberturas) e aberturas marcados 'Sim' ao mesmo
    tempo.

    NB: 'Não' COM Destino no Template e' esperado agora (o Destino e' uma
    sugestao preenchida em TODAS as linhas), por isso nao gera mais issue."""
    issues: list[str] = []
    sim_por_hier = defaultdict(list)
    origem_sim: dict[tuple, bool] = {}
    for row in rows:
        if str(row.get("alocacao_hierarquia", "")).strip() != ALOCACAO_SIM:
            continue
        grupo = str(row.get("grupo", "")).strip()
        hier = normalize_text(str(row.get("hierarquia", "")))
        origem = normalize_text(str(row.get("origem", "")))
        origem_sim[(grupo, origem)] = True
        if hier and hier != origem:
            sim_por_hier[(grupo, hier)].append(str(row.get("origem", "")))
    for (grupo, hier), aberturas in sim_por_hier.items():
        if origem_sim.get((grupo, hier)):
            issues.append(
                f"Dupla contagem (Grupo {grupo}): totalizador da Hierarquia "
                f"'{hier}' e suas aberturas {aberturas} estao ambos com "
                f"'Alocação da Hierarquia = Sim'. Escolher apenas um nivel."
            )
    return issues


def _item_tem_valor(item: dict[str, Any]) -> bool:
    """True se a conta (ja mesclada) tem ao menos um valor de ano nao-vazio.
    Trata "0"/0 como valor presente; so ""/None contam como ausencia."""
    for v in (item.get("valores_por_ano") or {}).values():
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return True
    return False


def _item_tem_valor_relevante(item: dict[str, Any]) -> bool:
    """True se a conta tem ao menos um valor de ano NAO-ZERO (e nao-vazio).
    Usado para decidir ALOCACAO: alocar (Sim) uma conta zerada em todos os anos
    vira ruido e polui a Memoria Anterior futura -> deve ficar como contexto
    (Nao). Usa coerce_number (entende formato BR '1.234,56', '(12)', etc.); '0',
    '0,00' e vazios contam como SEM valor relevante."""
    for v in (item.get("valores_por_ano") or {}).values():
        num = coerce_number(v)
        if num is None:
            continue
        if isinstance(num, (int, float)):
            if num != 0:
                return True
        else:
            # texto presente mas nao numerico -> nao tratar como zero
            return True
    return False


def compute_rastreabilidade_coverage(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Estatisticas (read-only) sobre a Rastreabilidade FINAL (apos o merge),
    para o parecer:
      - contas_capturadas: nº de contas distintas;
      - cobertura_valores: matriz Alocacao(Sim/Nao) x tem-valor(algum ano),
        espelhando a conferencia manual 'D=Sim/Nao com/sem valor' (o usuario
        nao precisa mais pedir esse cruzamento);
      - totalizadores: {total, alocados_sim} (pais detectados e quantos foram
        promovidos a 'Sim').
    CAPTURA COMPLETA: contexto (Nao) tambem guarda valor de ano."""
    merged = merge_rows_for_rastreabilidade(rows)
    totalizador_origens = compute_totalizador_origens(merged)
    cov = {"sim_com_valor": 0, "sim_sem_valor": 0, "nao_com_valor": 0, "nao_sem_valor": 0}
    pais_total: set[tuple[str, str]] = set()
    pais_sim: set[tuple[str, str]] = set()
    sim_zeradas = 0
    for item in merged:
        aloc = normalize_alocacao_value(item.get("alocacao_hierarquia"))
        slot = ("sim" if aloc == ALOCACAO_SIM else "nao") + (
            "_com_valor" if _item_tem_valor(item) else "_sem_valor"
        )
        cov[slot] += 1
        if aloc == ALOCACAO_SIM and not _item_tem_valor_relevante(item):
            sim_zeradas += 1
        origem_norm = normalize_text(_strip_totalizador_suffix(item.get("origem", "")))
        if origem_norm and origem_norm in totalizador_origens:
            chave_pai = (str(item.get("grupo", "")).strip(), origem_norm)
            pais_total.add(chave_pai)
            if aloc == ALOCACAO_SIM:
                pais_sim.add(chave_pai)
    return {
        "contas_capturadas": len(merged),
        "cobertura_valores": cov,
        "sim_alocadas_zeradas": sim_zeradas,
        "totalizadores": {"total": len(pais_total), "alocados_sim": len(pais_sim)},
    }


def validate_year_value_coverage(rows: list[dict[str, Any]]) -> list[str]:
    """Guardrail read-only (parecer): CAPTURA COMPLETA dos valores. Sinaliza
    conta ALOCADA (Alocacao da Hierarquia=Sim) sem valor RELEVANTE (vazio ou ZERO
    em todos os anos): nao faz sentido alocar conta zerada -- polui a Memoria
    Anterior futura; deve ser contexto (Nao). Linhas de contexto sem valor NAO
    viram issue (podem ser decorativas), mas aparecem na matriz 'cobertura_valores'
    do parecer p/ revisao."""
    issues: list[str] = []
    for item in merge_rows_for_rastreabilidade(rows):
        if normalize_alocacao_value(item.get("alocacao_hierarquia")) != ALOCACAO_SIM:
            continue
        if not _item_tem_valor_relevante(item):
            issues.append(
                f"Origem '{item.get('origem')}': alocada (Alocacao da Hierarquia=Sim) "
                f"mas SEM valor relevante (vazio ou ZERO em todos os anos) -- nao alocar conta "
                f"zerada (vira ruido na Memoria Anterior futura); marcar como contexto (Nao)."
            )
    return issues


def validate_totalizer_promotion(rows: list[dict[str, Any]]) -> list[str]:
    """Guardrail read-only (parecer) de SUB-CONSOLIDACAO -- espelha a dor do
    teste: a alocacao ficou analitica demais (filhos atomizados alocados, pai
    como contexto). Quando um totalizador esta 'Nao' e tem >=
    TOTALIZADOR_PROMOCAO_MIN_FILHOS aberturas atomizadas TODAS 'Sim' para o
    MESMO destino, sugere PROMOVER o totalizador a 'Sim' e rebaixar as aberturas
    a contexto (Nao), reduzindo a fragmentacao (Regras §2, 'Quando preferir o
    totalizador'). NAO sugere quando os filhos vao para destinos DIFERENTES (ai
    o totalizador misturaria classificacoes -- esse caso e' tratado por
    validate_sibling_consistency)."""
    merged = merge_rows_for_rastreabilidade(rows)
    totalizador_origens = compute_totalizador_origens(merged)

    pai_aloc: dict[tuple[str, str], str] = {}
    pai_label: dict[tuple[str, str], str] = {}
    filhos_dest: dict[tuple[str, str], set[str]] = defaultdict(set)
    filhos_qtd: dict[tuple[str, str], int] = defaultdict(int)
    dest_label: dict[tuple[str, str], str] = {}

    for item in merged:
        grupo = str(item.get("grupo", "")).strip()
        origem_norm = normalize_text(_strip_totalizador_suffix(item.get("origem", "")))
        hier_norm = normalize_text(_strip_totalizador_suffix(item.get("hierarquia", "")))
        aloc = normalize_alocacao_value(item.get("alocacao_hierarquia"))

        # Papel 1: a propria linha e' um totalizador (pai de alguma abertura).
        if origem_norm and origem_norm in totalizador_origens:
            pai_aloc[(grupo, origem_norm)] = aloc
            pai_label.setdefault((grupo, origem_norm), str(item.get("origem", "")).strip())

        # Papel 2: a linha e' abertura (filho) alocada (Sim) de um pai.
        if hier_norm and hier_norm != origem_norm and aloc == ALOCACAO_SIM:
            destino = str(item.get("destino_template", "")).strip()
            if destino:
                key = (grupo, hier_norm)
                filhos_dest[key].add(normalize_text(destino))
                filhos_qtd[key] += 1
                dest_label.setdefault(key, destino)
                pai_label.setdefault(key, str(item.get("hierarquia", "")).strip())

    issues: list[str] = []
    for key, qtd in filhos_qtd.items():
        if (
            qtd >= TOTALIZADOR_PROMOCAO_MIN_FILHOS
            and len(filhos_dest[key]) == 1
            and pai_aloc.get(key) != ALOCACAO_SIM
        ):
            issues.append(
                f"Bloco '{pai_label.get(key, key[1])}' (Grupo {key[0] or '?'}): "
                f"{qtd} aberturas atomizadas alocadas (Alocacao=Sim) todas para "
                f"'{dest_label.get(key)}', com o totalizador como 'Nao'. Avaliar "
                f"PROMOVER o totalizador a 'Sim' e deixar as aberturas como "
                f"contexto (Nao), reduzindo a fragmentacao (Regras §2)."
            )
    return issues


def generate_summary(rows: list[dict[str, Any]], issues: list[str], years: list[str]) -> dict[str, Any]:
    tipos = defaultdict(int)
    anos = set(years)
    grupos = defaultdict(float)

    alocacao = defaultdict(int)
    for row in rows:
        tipos[str(row.get("tipo_mapeamento", ""))] += 1
        alocacao[str(row.get("alocacao_hierarquia", "")) or "(vazio)"] += 1
        if row.get("ano"):
            anos.add(str(row.get("ano", "")).strip())
        if row.get("destino_template"):
            grupos[str(row.get("grupo", "")).strip()] += try_float(row.get("valor"))

    return {
        "quantidade_linhas": len(rows),
        "quantidade_alocadas": sum(1 for r in rows if r.get("destino_template")),
        "anos": sorted([a for a in anos if a], key=sort_year_key),
        "tipos_mapeamento": dict(tipos),
        "alocacao_hierarquia": dict(alocacao),
        **compute_rastreabilidade_coverage(rows),
        "totais_por_grupo_alocado": dict(grupos),
        "issues": issues,
    }


# =========================================================
# Abas internas (snapshot + Base de dados) e Main
# =========================================================
def _protect_sheet(ws, password: str) -> None:
    """Protege APENAS esta planilha contra edicao (anti-edicao acidental). Nao usa
    protecao de estrutura do workbook, portanto as demais abas continuam
    totalmente editaveis. Lembrete: protecao .xlsx nao e criptografia."""
    if password:
        ws.protection.password = password
    ws.protection.sheet = True
    ws.protection.enable()


def add_initial_rastreabilidade_snapshot(wb, rast_ws, password: str) -> None:
    """Cria a aba `rastreabilidade_inicial`: copia congelada da Rastreabilidade ja
    preenchida, para servir de referencia (espelho) e detectar quais contas
    mudaram depois. Sai oculta (LOCK_SHEET_STATE) + protegida por senha. As demais
    abas nao sao tocadas (continuam editaveis/visiveis)."""
    if SNAPSHOT_SHEET_NAME in wb.sheetnames:
        del wb[SNAPSHOT_SHEET_NAME]
    snap = wb.copy_worksheet(rast_ws)
    snap.title = SNAPSHOT_SHEET_NAME
    snap.sheet_state = LOCK_SHEET_STATE
    _protect_sheet(snap, password)


def count_reference_pages(rows: list[dict[str, Any]]) -> int:
    """Conta as paginas DISTINTAS que de fato trouxeram dados (origens do OCR):
    a soma das paginas de BP + DRE. Le 'pagina_referencia' de cada row (pode vir
    como '3' ou '3, 4'), separa por virgula/ponto-virgula/barra e conta unicos."""
    pages: set[str] = set()
    for row in rows:
        raw = str(row.get("pagina_referencia", "") or "").strip()
        if not raw:
            continue
        for tok in re.split(r"[,;/]+", raw):
            tok = tok.strip()
            if tok:
                pages.add(tok)
    return len(pages)


def normalize_unidade_medida(raw: Any) -> str:
    """Normaliza a unidade para apenas 'Mil', 'MM' ou 'Bi' (aba Base de dados). Mapeia
    variantes comuns; se nao reconhecer, devolve o texto original trimado."""
    key = normalize_text(raw)
    if not key:
        return ""
    if key in {"mil", "milhar", "milhares", "thousand", "thousands"}:
        return "Mil"
    if key in {"mm", "milhao", "milhoes", "mi", "mio", "million", "millions"}:
        return "MM"
    if key in {"bi", "bilhao", "bilhoes", "billion", "billions", "bn"}:
        return "Bi"
    return str(raw).strip()


def normalize_modificacao_valores(raw: Any) -> str:
    """Normaliza a 'Modificacao base de Valores' para 'Não', 'x1.000' ou '/1.000'.
    Default 'Não' (nao houve modificacao)."""
    s = str(raw or "").strip().lower()
    if not s or s in {"nao", "não", "n", "no", "none", "sem", "manter", "0"}:
        return "Não"
    if "/" in s or "div" in s or "÷" in s:
        return "/1.000"
    if "x" in s or "*" in s or "mult" in s:
        return "x1.000"
    return "Não"


def normalize_complexidade(raw: Any) -> str:
    """Normaliza o nivel de complexidade para 'Baixo', 'Médio' ou 'Alto'."""
    key = normalize_text(raw)
    if not key:
        return ""
    if key.startswith("baix") or key == "low":
        return "Baixo"
    if key.startswith("med") or key == "medium":
        return "Médio"
    if key.startswith("alt") or key == "high":
        return "Alto"
    return str(raw).strip()


def normalize_formato_auditado(raw: Any) -> str:
    """Normaliza a identificacao de auditoria para 'Sim'/'Não' (aba Base de dados).
    Reconhece 'auditado'/'nao auditado' e variantes Sim/Nao."""
    key = normalize_text(raw)
    if not key:
        return ""
    neg = any(t in key for t in ("nao", "sem", "unaud", "false"))
    if "auditad" in key or "audit" in key:
        return "Não" if neg else "Sim"
    if key in {"sim", "s", "yes", "true", "1"}:
        return "Sim"
    if key in {"nao", "n", "no", "false", "0"}:
        return "Não"
    return str(raw).strip()


def _format_clock(raw: Any) -> str:
    """Formata um horario para 'dd/MM/AAAA HH:MM:SS'. Tolera ISO
    ('2026-06-17T13:48:14'), aspas em volta, espaco como separador, sufixo 'Z' e
    offset com/sem ':'. Se nao reconhecer, devolve o texto original trimado."""
    s = str(raw or "").strip().strip('"').strip("'").strip()
    if not s:
        return ""
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M:%S")
    except ValueError:
        pass
    for fmt in (
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y %H:%M:%S")
        except ValueError:
            continue
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})[T ](\d{2}):(\d{2})(?::(\d{2}))?", s)
    if m:
        y, mo, d, h, mi, se = m.groups()
        return f"{d}/{mo}/{y} {h}:{mi}:{se or '00'}"
    return s


def _resolve_inicio_gpt(args) -> str:
    """Horario do UPLOAD do arquivo pelo usuario (log da GPT) para a aba Base de dados.
    Prioridade: --tempo-inicio-gpt > --start-time > --start-time-file."""
    raw = (getattr(args, "tempo_inicio_gpt", "") or "").strip()
    if not raw:
        raw = (getattr(args, "start_time", "") or "").strip()
    if not raw:
        stf = (getattr(args, "start_time_file", "") or "").strip()
        if stf:
            try:
                raw = Path(stf).read_text(encoding="utf-8").strip()
            except OSError:
                raw = ""
    return _format_clock(raw)


def build_base_dados_pairs(args, rows: list[dict[str, Any]], years: list[str], output_path: Path) -> list[tuple[str, Any]]:
    """Monta os pares (Cabecalho, Valor) da aba Base de dados. O layout final e' em
    COLUNAS (cabecalhos na linha 1, valores na linha 2 - ver fill_base_dados_sheet).
    Combina respostas confirmadas no chat + analises da GPT (auditoria,
    modificacao de valores, complexidade, horarios de inicio/fim) + metadados
    automaticos (empresa, CNPJ, anos, data/hora de geracao)."""
    paginas_ref = (getattr(args, "paginas_referencia", "") or "").strip()
    if not paginas_ref:
        _n = count_reference_pages(rows)
        paginas_ref = str(_n) if _n else ""
    tempo_final = _format_clock(getattr(args, "tempo_final_gpt", "")) or datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    return [
        ("Versão do GPT", (getattr(args, "versao_gpt", "") or VERSAO_GPT)),
        ("Modelo do arquivo", args.modelo),
        ("Formato Auditado", normalize_formato_auditado(getattr(args, "formato_auditado", ""))),
        ("Unidade de medida", normalize_unidade_medida(getattr(args, "unidade_medida", ""))),
        ("Moeda", getattr(args, "moeda", "")),
        ("Modificação base de Valores", normalize_modificacao_valores(getattr(args, "modificacao_valores", ""))),
        ("Páginas do input", args.paginas_input),
        ("Páginas de referência (BP+DRE)", paginas_ref),
        ("Nível de complexidade de alocação e planilhamento", normalize_complexidade(getattr(args, "complexidade", ""))),
        ("Tempo de Início (GPT)", _resolve_inicio_gpt(args)),
        ("Tempo Final (GPT)", tempo_final),
        ("Empresa", args.company_name),
        ("CNPJ", args.cnpj),
        ("Anos identificados", ", ".join(years)),
        ("Data/hora de geração", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Arquivo gerado", output_path.name),
    ]


def fill_base_dados_sheet(wb, config_pairs: list[tuple[str, Any]], password: str) -> None:
    """Preenche a aba `Base de dados` (que JA EXISTE no template novo, layout de
    COLUNAS: cabecalhos na linha 1, valores na linha 2). Casa cada par
    (campo, valor) com a coluna cujo cabecalho bate por NOME (normalizado) e
    escreve o valor na linha 2 -- robusto a reordenacao e a colunas extras do
    template (Matricula, Grupo, Segmento, # periodos, # alteracoes... ficam
    intactas, para preenchimento posterior). Preserva os cabecalhos/formatacao
    do template (so escreve a linha 2). Sai oculta (LOCK_SHEET_STATE) +
    protegida por senha, como a antiga `config`.
    Fallback: se a aba nao existir (template antigo), cria em layout de colunas."""
    alvo = BASE_DADOS_SHEET_NAME.strip().lower()
    ws = next((c for c in wb.worksheets if c.title.strip().lower() == alvo), None)

    if ws is not None:
        # Mapa cabecalho-normalizado -> coluna (linha 1, vinda do template).
        header_to_col: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            key = normalize_text(ws.cell(row=1, column=col).value)
            if key and key not in header_to_col:
                header_to_col[key] = col
        for campo, valor in config_pairs:
            col = header_to_col.get(normalize_text(campo))
            if not col:
                continue  # coluna nao existe no template -> ignora sem quebrar
            cell = ws.cell(row=2, column=col)
            cell.value = valor if isinstance(valor, (int, float)) else ("" if valor is None else str(valor))
    else:
        # Template antigo sem a aba: cria em layout de colunas (compatibilidade).
        ws = wb.create_sheet(title=BASE_DADOS_SHEET_NAME)
        for idx, (campo, valor) in enumerate(config_pairs, start=1):
            header = ws.cell(row=1, column=idx)
            header.value = str(campo)
            header.font = Font(bold=True, color="FFFFFFFF")
            header.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
            value_cell = ws.cell(row=2, column=idx)
            value_cell.value = valor if isinstance(valor, (int, float)) else ("" if valor is None else str(valor))
            ws.column_dimensions[get_column_letter(idx)].width = max(14, min(48, len(str(campo)) + 4))
        ws.freeze_panes = "A2"

    ws.sheet_state = LOCK_SHEET_STATE
    _protect_sheet(ws, password)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Preenche o template contabil preservando a estrutura original."
    )
    parser.add_argument("--template", required=True, help="Caminho do template Excel")
    parser.add_argument("--dictionary", required=True, help="Caminho do Dicionario de Contas.xlsx")
    parser.add_argument("--input-json", required=True, help="JSON com lancamentos estruturados do OCR")
    parser.add_argument("--output", required=True, help="Caminho do arquivo Excel final")
    parser.add_argument("--summary-json", default="", help="Resumo opcional em JSON")
    parser.add_argument(
        "--shadow-company",
        default="",
        help="Nome da empresa em B2 ou índice da aba Shadow quando houver múltiplas Shadows válidas",
    )
    parser.add_argument(
        "--company-name",
        default="",
        help="Nome da empresa para compor o nome padronizado do arquivo (OUTPUT_NAME_PATTERN)",
    )
    parser.add_argument(
        "--cnpj",
        default="",
        help="CNPJ da empresa para compor o nome padronizado do arquivo (so digitos sao usados)",
    )
    parser.add_argument(
        "--auto-name",
        action="store_true",
        help="Gera o nome do arquivo via OUTPUT_NAME_PATTERN, mantendo a pasta de --output",
    )
    parser.add_argument(
        "--modelo",
        default="",
        help="Modelo/visao confirmada no chat (ex.: Consolidado, Saldo Anterior) -> aba Base de dados",
    )
    parser.add_argument(
        "--paginas-input",
        default="",
        help="Numero de paginas do arquivo de input -> aba Base de dados",
    )
    parser.add_argument(
        "--paginas-referencia",
        default="",
        help="Qtd de paginas que de fato trouxeram BP+DRE; se vazio, o script conta as paginas distintas das origens -> aba Base de dados",
    )
    parser.add_argument(
        "--versao-gpt",
        default=VERSAO_GPT,
        help="Versao do GPT/knowledge (CustomGPT) -> aba Base de dados",
    )
    parser.add_argument(
        "--unidade-medida",
        default="",
        help="Unidade de medida confirmada no chat (ex.: Mil, MM, BI) -> aba Base de dados",
    )
    parser.add_argument(
        "--moeda",
        default="",
        help="Moeda confirmada no chat (ex.: BRL, US, EUR) -> aba Base de dados",
    )
    parser.add_argument(
        "--lock-password",
        default="",
        help="Senha anti-edicao das abas internas; vazio usa LOCK_PASSWORD do script",
    )
    parser.add_argument(
        "--start-time",
        default="",
        help="Horario ISO de inicio do fluxo (capturado pela GPT, ex.: 2026-06-02T09:40:00); o script calcula o tempo total input->output da aba Base de dados",
    )
    parser.add_argument(
        "--start-time-file",
        default="",
        help="Arquivo com o horario ISO de inicio (alternativa robusta a --start-time; sobrevive a reinicio de kernel)",
    )
    parser.add_argument(
        "--tempo-inicio-gpt",
        default="",
        help="Horario do UPLOAD do arquivo pelo usuario, logado pela GPT (ISO ou texto) -> 'Tempo de Inicio (GPT)' na aba Base de dados",
    )
    parser.add_argument(
        "--tempo-final-gpt",
        default="",
        help="Horario em que a GPT disponibilizou o arquivo no chat (ISO ou texto) -> 'Tempo Final (GPT)'; se vazio usa o instante de geracao",
    )
    parser.add_argument(
        "--modificacao-valores",
        default="",
        help="Modificacao base de valores aplicada pela GPT: Nao / x1.000 / /1.000 -> aba Base de dados",
    )
    parser.add_argument(
        "--complexidade",
        default="",
        help="Nivel de complexidade de alocacao e planilhamento avaliado pela GPT: Baixo/Medio/Alto -> aba Base de dados",
    )
    parser.add_argument(
        "--formato-auditado",
        default="",
        help="Identificacao da GPT se o arquivo de input era auditado: Sim/Nao -> aba Base de dados",
    )
    args = parser.parse_args()

    template_path = Path(args.template)
    dictionary_path = Path(args.dictionary)
    input_json = Path(args.input_json)
    output_path = Path(args.output)

    if not template_path.exists():
        raise FileNotFoundError(f"Template nao encontrado: {template_path}")
    if not dictionary_path.exists():
        raise FileNotFoundError(f"Dicionario nao encontrado: {dictionary_path}")
    if not input_json.exists():
        raise FileNotFoundError(f"JSON de entrada nao encontrado: {input_json}")

    _pipeline_start = time.perf_counter()
    _t = time.perf_counter()
    # Nome padronizado do arquivo de saida (opcional, via OUTPUT_NAME_PATTERN).
    # Acionado quando --company-name/--cnpj forem fornecidos ou --auto-name usado.
    if args.auto_name or args.company_name or args.cnpj:
        _ext = output_path.suffix if output_path.suffix.lower() in (".xlsx", ".xlsm") else (template_path.suffix or ".xlsx")
        output_path = output_path.parent / build_output_filename(
            args.company_name, args.cnpj, compute_years(normalize_rows(load_rows(input_json))), ext=_ext
        )
        print(f"[info] nome do arquivo de saida: {output_path.name}", file=sys.stderr)

    shutil.copy2(template_path, output_path)
    print(f"[timing] shutil.copy2(template->output): {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    _t = time.perf_counter()
    rows = normalize_rows(load_rows(input_json))
    dictionary_rows = load_dictionary(dictionary_path)
    print(f"[timing] load_rows + load_dictionary: {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    _t = time.perf_counter()
    wb = open_workbook(output_path)
    shadow_ws, rast_ws, listas_ws = validate_template(wb, args.shadow_company)
    print(f"[timing] open_workbook + validate_template: {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    template_index = build_template_account_index(shadow_ws)
    duplicates = find_duplicate_destination_names(template_index)

    _t = time.perf_counter()
    rows = [apply_special_classification_rules(r) for r in rows]

    rows = apply_shadow_mapping(rows, shadow_ws)
    rows = [apply_special_classification_rules(r) for r in rows]
    rows = enforce_template_classification(rows, template_index)

    rows = apply_dictionary_mapping(rows, dictionary_rows)
    rows = [apply_special_classification_rules(r) for r in rows]
    rows = enforce_template_classification(rows, template_index)

    for row in rows:
        # Alocação da Hierarquia (Sim/Não): julgamento AUTORITATIVO do modelo.
        # O Destino no Template e' preenchido em TODAS as linhas (sugestao p/
        # edicao manual); somente as linhas 'Sim' contam (Chave/Chave Destino).
        row["alocacao_hierarquia"] = normalize_alocacao_value(
            row.get("alocacao_hierarquia")
        )
        if row["alocacao_hierarquia"] == ALOCACAO_NAO:
            # Contexto: Destino e' apenas sugestao; Tipo de Mapeamento so faz
            # sentido para linhas alocadas (Sim).
            row["tipo_mapeamento"] = ""
        elif str(row.get("tipo_mapeamento", "")).strip() in ("", "Referência"):
            # Nivel escolhido (Sim) sem tipo informado -> Julgamental.
            row["tipo_mapeamento"] = "Julgamental"
        normalized = apply_special_classification_rules(row)
        row["grupo"] = normalized["grupo"]
        row["sub_categoria"] = normalized["sub_categoria"]

    rows = enforce_template_classification(rows, template_index)
    print(f"[timing] mappings + classification: {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    # Guardrail: rotula como 'Dicionário' as linhas pre-alocadas pelo modelo
    # cujo destino coincide com o dicionario (corrige o 'tudo Julgamental').
    rows, _dic_relabel = annotate_dictionary_source(rows, dictionary_rows)
    if _dic_relabel:
        print(f"[guardrail] {_dic_relabel} linha(s) rotulada(s) como Dicionário", file=sys.stderr)

    _t = time.perf_counter()
    shadow_protected_before = snapshot_protected_ranges(shadow_ws)
    print(f"[timing] snapshot_protected_ranges: {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    # Calcular os anos uma unica vez. Usados tanto na Rastreabilidade
    # (cabecalhos C1/D1/E1 e colunas de valor) quanto na Shadow.
    years = compute_years(rows)

    _t = time.perf_counter()
    start_row, end_row = append_rastreabilidade(rast_ws, rows, years, shadow_ws)
    propagate_formulas_down(rast_ws, start_row, end_row)
    fill_tracking_formulas(rast_ws, start_row, end_row)
    validate_tracking_columns_position(rast_ws)
    print(f"[timing] rastreabilidade writes (append+propagate+fill): {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    _t = time.perf_counter()
    set_year_headers(shadow_ws, years)
    update_shadow_memoria_atual(shadow_ws, rows)

    # apply_listas_formulas(listas_ws)
    # ^ desativada por escolha do usuario (Apr/2026): as formulas A2/C2/E2 da
    # Listas sao mantidas diretamente no template_plano_de_contas.xlsx. O script
    # NAO escreve formulas nem limpa A3+/C3+/E3+, preservando o que o usuario
    # gravou no template. Os pos-processamentos apply_dynamic_array_metadata e
    # apply_workbook_calc_features (chamados apos wb.save) continuam executando
    # para reaplicar cm="1" + metadata.xml + calcFeatures, que o openpyxl
    # destroi no ciclo load->save.
    apply_shadow_data_validations(shadow_ws)
    print(f"[timing] shadow updates (year+memoria+validations): {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    for ws in wb.worksheets:
        try:
            ws.sheet_view.showFormulas = False
        except Exception:
            pass

    _t = time.perf_counter()
    issues = []
    issues.extend(validate_subcategoria(rows))
    issues.extend(validate_duplicate_allocation(rows))
    issues.extend(validate_group_swaps(rows, dictionary_rows))
    issues.extend(validate_template_destination_structure(rows, template_index))
    issues.extend(validate_group_is_structurally_correct(rows, template_index))
    issues.extend(validate_tracking_formulas(rast_ws, start_row, end_row))
    issues.extend(validate_shadow_memory_structure(shadow_ws))
    issues.extend(validate_listas_formulas(listas_ws))
    issues.extend(validate_pl_specific_accounts(rows))
    issues.extend(validate_dre_subcategory(rows))
    issues.extend(validate_sibling_consistency(rows))
    issues.extend(validate_gpt_vs_dictionary(rows, dictionary_rows))
    issues.extend(validate_alocacao_consistency(rows))
    issues.extend(validate_totalizer_promotion(rows))
    issues.extend(validate_year_value_coverage(rows))
    print(f"[timing] all validate_* checks: {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    if duplicates:
        for nome, items in duplicates.items():
            issues.append(f"Template contém destino homônimo: {nome} -> {items}")

    # Capitalizar nomes das sheets canonicas para que o Excel exiba as
    # referencias `Rastreabilidade!`, `Shadow!`, `Listas!` (display case-
    # preserving das sheets reais) e nao as minusculas originais do template.
    _t = time.perf_counter()
    normalize_sheet_names(wb)
    print(f"[timing] normalize_sheet_names: {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    # Bloco active sheet=Listas + selection A2 desativado (Apr/2026): com as
    # formulas mantidas no template, preservar a active sheet / selection que o
    # usuario configurou. Se desejar restaurar, reverter este bloco.

    # Abas internas: snapshot congelado da Rastreabilidade + aba Base de dados
    # (Campo/Valor). Ambas saem ocultas (LOCK_SHEET_STATE) + protegidas por
    # senha; as demais abas seguem editaveis/visiveis (sem trava de estrutura).
    _lock_pw = args.lock_password or LOCK_PASSWORD
    add_initial_rastreabilidade_snapshot(wb, rast_ws, _lock_pw)
    fill_base_dados_sheet(wb, build_base_dados_pairs(args, rows, years, output_path), _lock_pw)

    _t = time.perf_counter()
    wb.save(output_path)
    print(f"[timing] wb.save: {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    _t = time.perf_counter()
    apply_dynamic_array_artifacts(output_path, template_path)
    verify_shadow_integrity(shadow_protected_before, output_path, shadow_ws.title)
    print(f"[timing] verify_shadow_integrity: {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    _t = time.perf_counter()
    verify_dynamic_array_artifacts(output_path, template_path)
    print(f"[timing] verify_dynamic_array_artifacts: {time.perf_counter() - _t:.2f}s", file=sys.stderr)

    summary = generate_summary(rows, issues, years)

    if args.summary_json:
        Path(args.summary_json).write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    print(f"[timing] PIPELINE TOTAL: {time.perf_counter() - _pipeline_start:.2f}s", file=sys.stderr)
    print(f"Arquivo gerado com sucesso: {output_path}")
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())